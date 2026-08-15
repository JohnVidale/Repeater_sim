#!/usr/bin/env python3
"""Estimate fixed-depth common lat/lon for repeater pairs from direct-P picks.

The fitter is intended for the South Sandwich repeater-pair workbook.  For each
configured pair it:

* keeps the workbook pair depth fixed,
* makes automatic direct-P AIC picks for both events,
* solves a common pair latitude/longitude plus separate origin-time offsets for
  the two events, and
* reports both a robust soft-L1 fit and a median-absolute-residual fit.

The median fit is preferred for the workbook update because a small number of
mispicked phases should not be allowed to drag the location.
"""

from __future__ import annotations

import argparse
import csv
import json
import math
import warnings
from dataclasses import dataclass
from datetime import UTC, datetime
from pathlib import Path
from typing import Any

import numpy as np
import openpyxl
from obspy import UTCDateTime, read
from obspy.geodetics import gps2dist_azimuth, locations2degrees
from obspy.signal.trigger import aic_simple
from obspy.taup import TauPyModel
from scipy.optimize import differential_evolution, least_squares, minimize


ROOT = Path(__file__).resolve().parent
CONFIG_PATH = ROOT / "analysis_config.json"

FILTER_HZ = (0.7, 4.0)
SEARCH_SECONDS = (-6.0, 8.0)
MIN_DISTANCE_DEGREES = 15.0
MAX_DISTANCE_DEGREES = 98.0
MIN_SNR = 2.5
MAX_PICK_OFFSET_SECONDS = 5.5
MIN_TOTAL_ACCEPTED_PICKS = 10
MIN_ACCEPTED_PICKS_PER_EVENT = 4
FINITE_DIFFERENCE_KM = 1.0
MAX_LINEAR_STEP_KM = 25.0
MAX_ITERATIONS = 3
EARTH_KM_PER_DEGREE = 111.195


@dataclass
class EventInfo:
    event_id: int
    origin: UTCDateTime


@dataclass
class PairInfo:
    label: str
    event1: EventInfo
    event2: EventInfo
    latitude: float
    longitude: float
    depth_km: float


@dataclass
class Pick:
    pair_label: str
    event_id: int
    station_id: str
    path: Path
    station_latitude: float
    station_longitude: float
    distance_degrees: float
    predicted_epoch: float
    pick_epoch: float
    pick_offset_seconds: float
    snr: float
    uncertainty_seconds: float
    accepted: bool
    rejection_reason: str


def normalized_header(row: tuple[Any, ...]) -> dict[str, int]:
    return {
        str(value).strip().lower(): index
        for index, value in enumerate(row)
        if value is not None and str(value).strip()
    }


def read_config(path: Path) -> dict[str, Any]:
    return json.loads(path.read_text(encoding="utf-8"))


def workbook_path(config: dict[str, Any]) -> Path:
    return Path(config.get("time_shift_workbook") or config["catalog_path"])


def load_pairs(config: dict[str, Any], labels: list[str]) -> list[PairInfo]:
    workbook = openpyxl.load_workbook(
        workbook_path(config), read_only=True, data_only=True
    )
    try:
        event_sheet = workbook["events"]
        event_rows = event_sheet.iter_rows(values_only=True)
        event_header = normalized_header(next(event_rows))
        events: dict[int, EventInfo] = {}
        for row in event_rows:
            if row[event_header["index"]] is None:
                continue
            event_id = int(row[event_header["index"]])
            events[event_id] = EventInfo(
                event_id=event_id,
                origin=UTCDateTime(str(row[event_header["time"]])),
            )

        pair_sheet = workbook["pairs"]
        pair_rows = pair_sheet.iter_rows(values_only=True)
        pair_header = normalized_header(next(pair_rows))
        selected = set(labels)
        pairs: list[PairInfo] = []
        for row in pair_rows:
            label_value = row[pair_header["label"]]
            if label_value is None:
                continue
            label = str(label_value).strip()
            if label not in selected:
                continue
            event1_id = int(row[pair_header["index1"]])
            event2_id = int(row[pair_header["index2"]])
            pairs.append(
                PairInfo(
                    label=label,
                    event1=events[event1_id],
                    event2=events[event2_id],
                    latitude=float(row[pair_header["lat"]]),
                    longitude=float(row[pair_header["lon"]]),
                    depth_km=float(row[pair_header["depth"]]),
                )
            )
    finally:
        workbook.close()
    found = {pair.label for pair in pairs}
    missing = [label for label in labels if label not in found]
    if missing:
        raise RuntimeError(f"Pairs not found in workbook: {', '.join(missing)}")
    return sorted(pairs, key=lambda pair: labels.index(pair.label))


def event_directory(waveform_root: Path, origin: UTCDateTime) -> Path:
    key = origin.strftime("%Y%m%d_%H%M%S")
    event_dir = waveform_root / key
    if event_dir.is_dir():
        return event_dir
    matches = sorted(path for path in waveform_root.glob(f"{key}*") if path.is_dir())
    if len(matches) == 1:
        return matches[0]
    raise RuntimeError(f"Could not resolve waveform directory for {origin}")


def direct_p_time(
    model: TauPyModel,
    latitude: float,
    longitude: float,
    depth_km: float,
    station_latitude: float,
    station_longitude: float,
) -> tuple[float, float] | None:
    distance = float(
        locations2degrees(latitude, longitude, station_latitude, station_longitude)
    )
    arrivals = model.get_travel_times(
        source_depth_in_km=float(depth_km),
        distance_in_degree=distance,
        phase_list=["P"],
    )
    if not arrivals:
        return None
    return float(arrivals[0].time), distance


def uncertainty_from_snr(snr: float) -> float:
    if snr >= 10.0:
        return 0.20
    if snr >= 6.0:
        return 0.35
    if snr >= 4.0:
        return 0.55
    return 0.85


def pick_trace(
    path: Path,
    pair_label: str,
    event: EventInfo,
    model: TauPyModel,
    latitude: float,
    longitude: float,
    depth_km: float,
) -> Pick | None:
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        try:
            trace = read(str(path))[0]
        except Exception:
            return None
    if str(trace.stats.channel).upper() != "BHZ":
        return None
    sac = trace.stats.sac
    try:
        station_latitude = float(sac.stla)
        station_longitude = float(sac.stlo)
    except (AttributeError, TypeError, ValueError):
        return None
    prediction = direct_p_time(
        model,
        latitude,
        longitude,
        depth_km,
        station_latitude,
        station_longitude,
    )
    if prediction is None:
        return None
    travel_time, distance_degrees = prediction
    if not MIN_DISTANCE_DEGREES <= distance_degrees <= MAX_DISTANCE_DEGREES:
        return None
    predicted_epoch = float(event.origin) + travel_time

    processed = trace.copy()
    processed.detrend("demean")
    processed.detrend("linear")
    processed.taper(max_percentage=0.05, max_length=2.0, type="cosine")
    processed.filter(
        "bandpass",
        freqmin=FILTER_HZ[0],
        freqmax=FILTER_HZ[1],
        corners=4,
        zerophase=True,
    )
    fs = float(processed.stats.sampling_rate)
    start_epoch = float(processed.stats.starttime)
    data = np.asarray(processed.data, dtype=float)
    relative = start_epoch + np.arange(len(data)) / fs - predicted_epoch
    indices = np.flatnonzero(
        (relative >= SEARCH_SECONDS[0]) & (relative < SEARCH_SECONDS[1])
    )
    if len(indices) < int(8.0 * fs):
        return None
    segment = data[indices]
    characteristic = np.asarray(aic_simple(segment), dtype=float)
    edge = max(1, int(round(0.5 * fs)))
    interior = characteristic[edge:-edge]
    if not len(interior) or not np.any(np.isfinite(interior)):
        return None
    local_index = edge + int(np.nanargmin(interior))
    pick_index = int(indices[local_index])
    pick_epoch = start_epoch + pick_index / fs
    pick_offset = pick_epoch - predicted_epoch
    edge_margin = min(
        pick_offset - SEARCH_SECONDS[0], SEARCH_SECONDS[1] - pick_offset
    )

    pick_relative = start_epoch + np.arange(len(data)) / fs - pick_epoch
    noise = data[(pick_relative >= -8.0) & (pick_relative < -2.0)]
    signal = data[(pick_relative >= 0.0) & (pick_relative < 4.0)]
    if len(noise) < int(3.0 * fs) or len(signal) < int(2.0 * fs):
        return None
    noise_rms = float(np.sqrt(np.mean(np.square(noise))))
    signal_rms = float(np.sqrt(np.mean(np.square(signal))))
    snr = signal_rms / noise_rms if noise_rms > 0.0 else math.inf
    reasons = []
    if snr < MIN_SNR:
        reasons.append(f"snr<{MIN_SNR:g}")
    if edge_margin < 0.75:
        reasons.append("pick_near_search_edge")
    if abs(pick_offset) > MAX_PICK_OFFSET_SECONDS:
        reasons.append(f"pick_offset>{MAX_PICK_OFFSET_SECONDS:g}s")
    return Pick(
        pair_label=pair_label,
        event_id=event.event_id,
        station_id=f"{trace.stats.network}.{trace.stats.station}",
        path=path,
        station_latitude=station_latitude,
        station_longitude=station_longitude,
        distance_degrees=distance_degrees,
        predicted_epoch=predicted_epoch,
        pick_epoch=pick_epoch,
        pick_offset_seconds=pick_offset,
        snr=snr,
        uncertainty_seconds=uncertainty_from_snr(snr),
        accepted=not reasons,
        rejection_reason=";".join(reasons),
    )


def collect_pair_picks(
    pair: PairInfo, waveform_root: Path, model: TauPyModel, excluded_stations: set[str] | None = None
) -> list[Pick]:
    picks: list[Pick] = []
    for event in (pair.event1, pair.event2):
        directory = event_directory(waveform_root, event.origin)
        for path in sorted(directory.rglob("*")):
            if not path.is_file():
                continue
            pick = pick_trace(
                path,
                pair.label,
                event,
                model,
                pair.latitude,
                pair.longitude,
                pair.depth_km,
            )
            if pick is not None and not (
                excluded_stations
                and pick.station_id.upper().split(".")[-1] in excluded_stations
            ):
                picks.append(pick)
    return picks


def shifted_location(latitude: float, longitude: float, east_km: float, north_km: float) -> tuple[float, float]:
    cosine_latitude = max(0.05, math.cos(math.radians(latitude)))
    return (
        latitude + north_km / EARTH_KM_PER_DEGREE,
        longitude + east_km / (EARTH_KM_PER_DEGREE * cosine_latitude),
    )


def travel_time_gradient(
    model: TauPyModel,
    pick: Pick,
    latitude: float,
    longitude: float,
    depth_km: float,
) -> tuple[float, float, float] | None:
    center = direct_p_time(
        model,
        latitude,
        longitude,
        depth_km,
        pick.station_latitude,
        pick.station_longitude,
    )
    if center is None:
        return None
    values = [float(center[0])]
    for east_km, north_km in (
        (FINITE_DIFFERENCE_KM, 0.0),
        (-FINITE_DIFFERENCE_KM, 0.0),
        (0.0, FINITE_DIFFERENCE_KM),
        (0.0, -FINITE_DIFFERENCE_KM),
    ):
        lat, lon = shifted_location(latitude, longitude, east_km, north_km)
        shifted = direct_p_time(
            model,
            lat,
            lon,
            depth_km,
            pick.station_latitude,
            pick.station_longitude,
        )
        if shifted is None:
            return None
        values.append(float(shifted[0]))
    gradient_east = (values[1] - values[2]) / (2.0 * FINITE_DIFFERENCE_KM)
    gradient_north = (values[3] - values[4]) / (2.0 * FINITE_DIFFERENCE_KM)
    return values[0], gradient_east, gradient_north


def design_at_location(
    model: TauPyModel,
    pair: PairInfo,
    picks: list[Pick],
    latitude: float,
    longitude: float,
) -> tuple[np.ndarray, np.ndarray, np.ndarray, list[Pick]]:
    rows: list[list[float]] = []
    residuals: list[float] = []
    weights: list[float] = []
    used: list[Pick] = []
    for pick in picks:
        gradient = travel_time_gradient(
            model, pick, latitude, longitude, pair.depth_km
        )
        if gradient is None:
            continue
        predicted, gradient_east, gradient_north = gradient
        origin = pair.event1.origin if pick.event_id == pair.event1.event_id else pair.event2.origin
        event1_column = 1.0 if pick.event_id == pair.event1.event_id else 0.0
        event2_column = 1.0 if pick.event_id == pair.event2.event_id else 0.0
        rows.append([gradient_east, gradient_north, event1_column, event2_column])
        residuals.append(float(pick.pick_epoch) - float(origin) - predicted)
        weights.append(max(0.05, float(pick.uncertainty_seconds)))
        used.append(pick)
    return (
        np.asarray(rows, dtype=float),
        np.asarray(residuals, dtype=float),
        np.asarray(weights, dtype=float),
        used,
    )


def median_offsets(
    design: np.ndarray, residuals: np.ndarray, event_ids: np.ndarray, east: float, north: float
) -> tuple[float, float, np.ndarray]:
    preliminary = residuals - design[:, :2] @ np.asarray([east, north], dtype=float)
    offset1 = float(np.median(preliminary[event_ids == 1]))
    offset2 = float(np.median(preliminary[event_ids == 2]))
    offsets = np.where(event_ids == 1, offset1, offset2)
    return offset1, offset2, offsets


def solve_linearized(
    design: np.ndarray,
    residuals: np.ndarray,
    weights: np.ndarray,
    used: list[Pick],
) -> dict[str, Any]:
    event_ids = np.asarray([1 if pick.event_id == used[0].event_id else 2 for pick in used])

    def soft_l1_residual(params: np.ndarray) -> np.ndarray:
        modeled = design @ params
        return (residuals - modeled) / weights

    soft = least_squares(
        soft_l1_residual,
        np.zeros(4, dtype=float),
        bounds=(
            [-MAX_LINEAR_STEP_KM, -MAX_LINEAR_STEP_KM, -20.0, -20.0],
            [MAX_LINEAR_STEP_KM, MAX_LINEAR_STEP_KM, 20.0, 20.0],
        ),
        loss="soft_l1",
        f_scale=2.0,
        max_nfev=200,
        xtol=1e-10,
        ftol=1e-10,
        gtol=1e-10,
    )

    def median_objective(horizontal: np.ndarray) -> float:
        east, north = map(float, horizontal)
        _, _, offsets = median_offsets(design, residuals, event_ids, east, north)
        modeled = design[:, :2] @ np.asarray([east, north]) + offsets
        return float(np.median(np.abs(residuals - modeled)))

    global_fit = differential_evolution(
        median_objective,
        bounds=[(-MAX_LINEAR_STEP_KM, MAX_LINEAR_STEP_KM)] * 2,
        seed=20260813,
        polish=False,
        tol=1e-8,
        updating="immediate",
        workers=1,
    )
    local_fit = minimize(
        median_objective,
        global_fit.x,
        method="Powell",
        bounds=[(-MAX_LINEAR_STEP_KM, MAX_LINEAR_STEP_KM)] * 2,
        options={"maxiter": 1000, "xtol": 1e-10, "ftol": 1e-12},
    )
    horizontal = local_fit.x if median_objective(local_fit.x) <= median_objective(global_fit.x) else global_fit.x
    east, north = map(float, horizontal)
    median_offset1, median_offset2, median_event_offsets = median_offsets(
        design, residuals, event_ids, east, north
    )
    median_modeled = design[:, :2] @ horizontal + median_event_offsets
    median_residuals = residuals - median_modeled
    soft_residuals = residuals - design @ soft.x
    return {
        "soft_l1": {
            "east_km": float(soft.x[0]),
            "north_km": float(soft.x[1]),
            "event1_origin_offset_s": float(soft.x[2]),
            "event2_origin_offset_s": float(soft.x[3]),
            "residuals": soft_residuals,
        },
        "median": {
            "east_km": east,
            "north_km": north,
            "event1_origin_offset_s": median_offset1,
            "event2_origin_offset_s": median_offset2,
            "residuals": median_residuals,
            "objective_median_abs_s": float(np.median(np.abs(median_residuals))),
        },
    }


def fit_pair(pair: PairInfo, picks: list[Pick], model: TauPyModel) -> dict[str, Any]:
    accepted = [pick for pick in picks if pick.accepted]
    counts = {
        pair.event1.event_id: sum(1 for pick in accepted if pick.event_id == pair.event1.event_id),
        pair.event2.event_id: sum(1 for pick in accepted if pick.event_id == pair.event2.event_id),
    }
    if (
        len(accepted) < MIN_TOTAL_ACCEPTED_PICKS
        or counts[pair.event1.event_id] < MIN_ACCEPTED_PICKS_PER_EVENT
        or counts[pair.event2.event_id] < MIN_ACCEPTED_PICKS_PER_EVENT
    ):
        return {
            "status": "failed_insufficient_picks",
            "accepted_picks": len(accepted),
            "event1_accepted_picks": counts[pair.event1.event_id],
            "event2_accepted_picks": counts[pair.event2.event_id],
        }

    latitude = pair.latitude
    longitude = pair.longitude
    final_used: list[Pick] = []
    final_solution: dict[str, Any] | None = None
    final_base_latitude = latitude
    final_base_longitude = longitude
    for _ in range(MAX_ITERATIONS):
        design, residuals, weights, used = design_at_location(
            model, pair, accepted, latitude, longitude
        )
        if len(used) < MIN_TOTAL_ACCEPTED_PICKS:
            return {
                "status": "failed_insufficient_gradient_rows",
                "accepted_picks": len(accepted),
                "gradient_rows": len(used),
            }
        solution = solve_linearized(design, residuals, weights, used)
        final_base_latitude = latitude
        final_base_longitude = longitude
        east = float(solution["median"]["east_km"])
        north = float(solution["median"]["north_km"])
        latitude, longitude = shifted_location(latitude, longitude, east, north)
        final_used = used
        final_solution = solution
        if math.hypot(east, north) < 0.05:
            break

    assert final_solution is not None
    median_horizontal = gps2dist_azimuth(pair.latitude, pair.longitude, latitude, longitude)[0] / 1000.0
    median_azimuth = gps2dist_azimuth(pair.latitude, pair.longitude, latitude, longitude)[1]
    median_east = median_horizontal * math.sin(math.radians(median_azimuth))
    median_north = median_horizontal * math.cos(math.radians(median_azimuth))

    soft = final_solution["soft_l1"]
    soft_lat, soft_lon = shifted_location(
        final_base_latitude,
        final_base_longitude,
        float(soft["east_km"]),
        float(soft["north_km"]),
    )
    soft_horizontal = gps2dist_azimuth(pair.latitude, pair.longitude, soft_lat, soft_lon)[0] / 1000.0
    soft_azimuth = gps2dist_azimuth(pair.latitude, pair.longitude, soft_lat, soft_lon)[1]
    soft_residuals = np.asarray(soft["residuals"], dtype=float)
    median_residuals = np.asarray(final_solution["median"]["residuals"], dtype=float)
    return {
        "status": "ok",
        "accepted_picks": len(accepted),
        "event1_accepted_picks": counts[pair.event1.event_id],
        "event2_accepted_picks": counts[pair.event2.event_id],
        "gradient_rows": len(final_used),
        "soft_l1_lat": soft_lat,
        "soft_l1_lon": soft_lon,
        "soft_l1_east_km": soft_horizontal * math.sin(math.radians(soft_azimuth)),
        "soft_l1_north_km": soft_horizontal * math.cos(math.radians(soft_azimuth)),
        "soft_l1_event1_origin_offset_s": float(soft["event1_origin_offset_s"]),
        "soft_l1_event2_origin_offset_s": float(soft["event2_origin_offset_s"]),
        "soft_l1_median_abs_residual_s": float(np.median(np.abs(soft_residuals))),
        "soft_l1_rms_residual_s": float(np.sqrt(np.mean(soft_residuals * soft_residuals))),
        "median_lat": latitude,
        "median_lon": longitude,
        "median_east_km": median_east,
        "median_north_km": median_north,
        "median_horizontal_km": median_horizontal,
        "median_event1_origin_offset_s": float(final_solution["median"]["event1_origin_offset_s"]),
        "median_event2_origin_offset_s": float(final_solution["median"]["event2_origin_offset_s"]),
        "median_abs_residual_s": float(np.median(np.abs(median_residuals))),
        "median_rms_residual_s": float(np.sqrt(np.mean(median_residuals * median_residuals))),
        "used_picks": final_used,
        "median_residuals": median_residuals,
    }


def write_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    fields = list(dict.fromkeys(key for row in rows for key in row))
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields)
        writer.writeheader()
        writer.writerows(rows)


def run(config_path: Path, output: Path, pair_labels: list[str] | None) -> None:
    config = read_config(config_path)
    labels = pair_labels or [str(label) for label in config["pairs"]]
    pairs = load_pairs(config, labels)
    waveform_root = Path(config["waveform_root"])
    model = TauPyModel(str(config.get("taup_model", "ak135")))
    excluded_stations = {
        str(value).strip().upper().split(".")[-1]
        for value in config.get("excluded_stations", [])
        if str(value).strip()
    }
    output.mkdir(parents=True, exist_ok=True)

    summary_rows: list[dict[str, Any]] = []
    residual_rows: list[dict[str, Any]] = []
    pick_rows: list[dict[str, Any]] = []
    for index, pair in enumerate(pairs, start=1):
        print(f"{index}/{len(pairs)} {pair.label}: picking direct P", flush=True)
        picks = collect_pair_picks(pair, waveform_root, model, excluded_stations)
        result = fit_pair(pair, picks, model)
        common = {
            "pair_label": pair.label,
            "event1": pair.event1.event_id,
            "event2": pair.event2.event_id,
            "initial_lat": pair.latitude,
            "initial_lon": pair.longitude,
            "fixed_depth_km": pair.depth_km,
            "status": result["status"],
        }
        if result["status"] == "ok":
            summary_rows.append(
                {
                    **common,
                    "accepted_picks": result["accepted_picks"],
                    "event1_accepted_picks": result["event1_accepted_picks"],
                    "event2_accepted_picks": result["event2_accepted_picks"],
                    "gradient_rows": result["gradient_rows"],
                    "preferred_method": "median",
                    "new_lat": result["median_lat"],
                    "new_lon": result["median_lon"],
                    "median_east_km": result["median_east_km"],
                    "median_north_km": result["median_north_km"],
                    "median_horizontal_km": result["median_horizontal_km"],
                    "median_event1_origin_offset_s": result["median_event1_origin_offset_s"],
                    "median_event2_origin_offset_s": result["median_event2_origin_offset_s"],
                    "median_abs_residual_s": result["median_abs_residual_s"],
                    "median_rms_residual_s": result["median_rms_residual_s"],
                    "soft_l1_lat": result["soft_l1_lat"],
                    "soft_l1_lon": result["soft_l1_lon"],
                    "soft_l1_east_km": result["soft_l1_east_km"],
                    "soft_l1_north_km": result["soft_l1_north_km"],
                    "soft_l1_event1_origin_offset_s": result["soft_l1_event1_origin_offset_s"],
                    "soft_l1_event2_origin_offset_s": result["soft_l1_event2_origin_offset_s"],
                    "soft_l1_median_abs_residual_s": result["soft_l1_median_abs_residual_s"],
                    "soft_l1_rms_residual_s": result["soft_l1_rms_residual_s"],
                }
            )
            for pick, residual in zip(result["used_picks"], result["median_residuals"]):
                residual_rows.append(
                    {
                        **common,
                        "station_id": pick.station_id,
                        "event_id": pick.event_id,
                        "station_latitude": pick.station_latitude,
                        "station_longitude": pick.station_longitude,
                        "distance_degrees_initial": pick.distance_degrees,
                        "snr": pick.snr,
                        "median_fit_residual_s": float(residual),
                        "waveform_path": str(pick.path),
                    }
                )
        else:
            summary_rows.append(
                {
                    **common,
                    "accepted_picks": result.get("accepted_picks", ""),
                    "event1_accepted_picks": result.get("event1_accepted_picks", ""),
                    "event2_accepted_picks": result.get("event2_accepted_picks", ""),
                    "gradient_rows": result.get("gradient_rows", ""),
                }
            )
        for pick in picks:
            pick_rows.append(
                {
                    **common,
                    "station_id": pick.station_id,
                    "event_id": pick.event_id,
                    "station_latitude": pick.station_latitude,
                    "station_longitude": pick.station_longitude,
                    "distance_degrees_initial": pick.distance_degrees,
                    "pick_time_utc": str(UTCDateTime(pick.pick_epoch)),
                    "pick_offset_from_initial_prediction_s": pick.pick_offset_seconds,
                    "snr": pick.snr,
                    "accepted": pick.accepted,
                    "rejection_reason": pick.rejection_reason,
                    "waveform_path": str(pick.path),
                }
            )
        print(
            f"{pair.label}: {result['status']} "
            f"accepted={result.get('accepted_picks', '')}",
            flush=True,
        )

    write_csv(output / "pair_fixed_depth_location_summary.csv", summary_rows)
    write_csv(output / "pair_fixed_depth_pick_residuals.csv", residual_rows)
    write_csv(output / "pair_fixed_depth_picks.csv", pick_rows)
    manifest = {
        "created_utc": datetime.now(UTC).isoformat(timespec="seconds").replace("+00:00", "Z"),
        "config_path": str(config_path),
        "workbook_path": str(workbook_path(config)),
        "pair_labels": labels,
        "method": {
            "phase": "direct P",
            "taup_model": str(config.get("taup_model", "ak135")),
            "filter_hz": FILTER_HZ,
            "search_seconds_relative_to_initial_prediction": SEARCH_SECONDS,
            "minimum_snr": MIN_SNR,
            "distance_range_degrees": [MIN_DISTANCE_DEGREES, MAX_DISTANCE_DEGREES],
            "fixed_depth_source": "pairs sheet depth",
            "preferred_location_fit": "median absolute residual",
            "comparison_fit": "least_squares soft_l1",
            "finite_difference_step_km": FINITE_DIFFERENCE_KM,
            "maximum_linearized_step_km_per_iteration": MAX_LINEAR_STEP_KM,
            "iterations": MAX_ITERATIONS,
        },
        "outputs": [
            "pair_fixed_depth_location_summary.csv",
            "pair_fixed_depth_pick_residuals.csv",
            "pair_fixed_depth_picks.csv",
        ],
    }
    (output / "manifest.json").write_text(json.dumps(manifest, indent=2) + "\n")
    print(output)


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--config", type=Path, default=CONFIG_PATH)
    parser.add_argument("--output", type=Path, default=None)
    parser.add_argument("--pairs", nargs="*", default=None)
    arguments = parser.parse_args()
    output = arguments.output
    if output is None:
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output = ROOT / "outputs" / f"fixed_depth_pair_locations_{stamp}"
    run(arguments.config, output, arguments.pairs)


if __name__ == "__main__":
    main()
