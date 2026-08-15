#!/usr/bin/env python3
"""Write preferred pair centroid and repeater-offset columns to ICevents_full.xlsx."""

from __future__ import annotations

import argparse
import csv
import shutil
from datetime import datetime
from pathlib import Path
from typing import Any

import openpyxl


DEFAULT_WORKBOOK = Path("/Users/jvidale/Documents/GitHub/Array_codes/Files/ICevents_full.xlsx")
DEFAULT_RELATIVE_LOCATIONS = Path(
    "outputs/multiphase_median_cc0.5_workbook_20260813_165113/"
    "median_absolute_relative_locations_no_pkikp.csv"
)


def normalized_header(row: tuple[Any, ...]) -> dict[str, int]:
    return {
        str(value).strip().lower(): index + 1
        for index, value in enumerate(row)
        if value is not None and str(value).strip()
    }


def read_relative_locations(path: Path) -> dict[str, dict[str, float | str | int]]:
    rows: dict[str, dict[str, float | str | int]] = {}
    with path.open(newline="", encoding="utf-8") as handle:
        for row in csv.DictReader(handle):
            pair = row["pair"].strip()
            rows[pair] = {
                "phase_set": row["phase_set"],
                "n": int(row["n"]),
                "event1": int(row["event1"]),
                "event2": int(row["event2"]),
                "event1_lat": float(row["event1_lat"]),
                "event1_lon": float(row["event1_lon"]),
                "event1_depth_km": float(row["event1_depth_km"]),
                "event2_lat": float(row["event2_lat"]),
                "event2_lon": float(row["event2_lon"]),
                "event2_depth_km": float(row["event2_depth_km"]),
                "delta_lat_degrees": float(row["delta_lat_degrees"]),
                "delta_lon_degrees": float(row["delta_lon_degrees"]),
                "depth_diff_km": float(row["depth_diff_km"]),
                "event2_minus_event1_origin_time_shift_s": float(
                    row["event2_minus_event1_origin_time_shift_s"]
                ),
                "east_km": float(row["east_km"]),
                "north_km": float(row["north_km"]),
                "horizontal_km": float(row["horizontal_km"]),
                "separation_3d_km": float(row["separation_3d_km"]),
                "median_abs_fit_residual_s": float(row["median_abs_fit_residual_s"]),
                "residual_rms_s": float(row["residual_rms_s"]),
            }
    return rows


def ensure_columns(sheet, names: list[str]) -> dict[str, int]:
    header = normalized_header(
        tuple(cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1)))
    )
    next_col = sheet.max_column + 1
    for name in names:
        key = name.lower()
        if key in header:
            continue
        sheet.cell(row=1, column=next_col).value = name
        header[key] = next_col
        next_col += 1
    return header


def write_offsets(workbook_path: Path, relative_locations_path: Path) -> Path:
    locations = read_relative_locations(relative_locations_path)
    backup_path = (
        workbook_path.parent
        / f"{workbook_path.stem}_before_pair_centroid_offsets_"
        f"{datetime.now().strftime('%Y%m%d_%H%M%S')}{workbook_path.suffix}"
    )
    shutil.copy2(workbook_path, backup_path)

    workbook = openpyxl.load_workbook(workbook_path)
    try:
        sheet = workbook["pairs"]
        columns = [
            "relative_location_phase_set",
            "relative_location_n",
            "centroid_lat",
            "centroid_lon",
            "centroid_depth_km",
            "event2_minus_event1_delta_lat",
            "event2_minus_event1_delta_lon",
            "event2_minus_event1_delta_depth_km",
            "event2_minus_event1_delta_time_s",
            "event2_minus_event1_3d_km",
            "relative_location_median_abs_residual_s",
            "relative_location_rms_residual_s",
        ]
        header = ensure_columns(sheet, columns)
        label_column = header["label"]
        depth_column = header["depth"]
        new_lat_column = header.get("new_lat")
        new_lon_column = header.get("new_lon")
        fallback_lat_column = header["lat"]
        fallback_lon_column = header["lon"]
        written = 0
        for row_index in range(2, sheet.max_row + 1):
            label_value = sheet.cell(row=row_index, column=label_column).value
            if label_value is None:
                continue
            label = str(label_value).strip()
            if label not in locations:
                continue
            row = locations[label]
            delta_lat = float(row["delta_lat_degrees"])
            delta_lon = float(row["delta_lon_degrees"])
            delta_depth = float(row["depth_diff_km"])
            delta_time = float(row["event2_minus_event1_origin_time_shift_s"])
            centroid_lat_cell = (
                sheet.cell(row=row_index, column=new_lat_column).value
                if new_lat_column is not None
                else None
            )
            centroid_lon_cell = (
                sheet.cell(row=row_index, column=new_lon_column).value
                if new_lon_column is not None
                else None
            )
            centroid_lat = (
                float(centroid_lat_cell)
                if centroid_lat_cell is not None
                else float(sheet.cell(row=row_index, column=fallback_lat_column).value)
            )
            centroid_lon = (
                float(centroid_lon_cell)
                if centroid_lon_cell is not None
                else float(sheet.cell(row=row_index, column=fallback_lon_column).value)
            )
            centroid_depth = float(sheet.cell(row=row_index, column=depth_column).value)
            values = {
                "relative_location_phase_set": row["phase_set"],
                "relative_location_n": row["n"],
                "centroid_lat": centroid_lat,
                "centroid_lon": centroid_lon,
                "centroid_depth_km": centroid_depth,
                "event2_minus_event1_delta_lat": delta_lat,
                "event2_minus_event1_delta_lon": delta_lon,
                "event2_minus_event1_delta_depth_km": delta_depth,
                "event2_minus_event1_delta_time_s": delta_time,
                "event2_minus_event1_3d_km": row["separation_3d_km"],
                "relative_location_median_abs_residual_s": row[
                    "median_abs_fit_residual_s"
                ],
                "relative_location_rms_residual_s": row["residual_rms_s"],
            }
            for name, value in values.items():
                cell = sheet.cell(row=row_index, column=header[name.lower()])
                cell.value = value
                if isinstance(value, float):
                    cell.number_format = "0.0000"
            written += 1
        workbook.save(workbook_path)
    finally:
        workbook.close()
    print(f"backup={backup_path}")
    print(f"written={written}")
    return backup_path


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--workbook", type=Path, default=DEFAULT_WORKBOOK)
    parser.add_argument("--relative-locations", type=Path, default=DEFAULT_RELATIVE_LOCATIONS)
    args = parser.parse_args()
    write_offsets(args.workbook, args.relative_locations)


if __name__ == "__main__":
    main()
