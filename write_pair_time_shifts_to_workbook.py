#!/usr/bin/env python3
"""Write computed pair median shifts into ICevents_full.xlsx."""

from __future__ import annotations

import argparse
import csv
import math
import shutil
from datetime import datetime
from pathlib import Path
from typing import Any

import openpyxl


def normalized_header(row: tuple[Any, ...]) -> dict[str, int]:
    return {
        str(value).strip().lower(): index + 1
        for index, value in enumerate(row)
        if value is not None and str(value).strip()
    }


def finite_float(value: Any) -> float | None:
    try:
        parsed = float(value)
    except (TypeError, ValueError):
        return None
    if not math.isfinite(parsed):
        return None
    return parsed


def read_computed_shifts(summary_path: Path) -> dict[str, float]:
    shifts: dict[str, float] = {}
    with summary_path.open(newline="", encoding="utf-8") as handle:
        for row in csv.DictReader(handle):
            label = str(row.get("pair_label", "")).strip()
            value = finite_float(row.get("computed_median_shift_seconds"))
            if not label or value is None:
                continue
            shifts[label] = value
    return shifts


def write_shifts(
    workbook_path: Path,
    median_summary_path: Path,
    *,
    only_missing: bool,
) -> Path:
    shifts = read_computed_shifts(median_summary_path)
    backup_path = (
        workbook_path.parent
        / f"{workbook_path.stem}_before_new_time_shifts_"
        f"{datetime.now().strftime('%Y%m%d_%H%M%S')}{workbook_path.suffix}"
    )
    shutil.copy2(workbook_path, backup_path)

    workbook = openpyxl.load_workbook(workbook_path)
    try:
        sheet = workbook["pairs"]
        header = normalized_header(
            tuple(cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1)))
        )
        label_column = header["label"]
        shift_column = header.get("new time shift")
        if shift_column is None:
            shift_column = sheet.max_column + 1
            sheet.cell(row=1, column=shift_column).value = "new time shift"
        written = 0
        skipped_existing = 0
        skipped_missing = 0
        for row_index in range(2, sheet.max_row + 1):
            label_value = sheet.cell(row=row_index, column=label_column).value
            if label_value is None:
                continue
            label = str(label_value).strip()
            if label not in shifts:
                continue
            current = finite_float(sheet.cell(row=row_index, column=shift_column).value)
            if only_missing and current is not None:
                skipped_existing += 1
                continue
            value = shifts[label]
            if not math.isfinite(value):
                skipped_missing += 1
                continue
            cell = sheet.cell(row=row_index, column=shift_column)
            cell.value = value
            cell.number_format = "0.0000"
            written += 1
        workbook.save(workbook_path)
    finally:
        workbook.close()

    print(f"backup={backup_path}")
    print(f"written={written}")
    print(f"skipped_existing={skipped_existing}")
    print(f"skipped_missing={skipped_missing}")
    return backup_path


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("workbook", type=Path)
    parser.add_argument("median_summary", type=Path)
    parser.add_argument(
        "--overwrite",
        action="store_true",
        help="Overwrite existing new time shift values instead of filling blanks only.",
    )
    args = parser.parse_args()
    write_shifts(args.workbook, args.median_summary, only_missing=not args.overwrite)


if __name__ == "__main__":
    main()
