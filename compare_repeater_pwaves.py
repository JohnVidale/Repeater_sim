#!/usr/bin/env python3
"""Compare aligned repeating-earthquake P waves with their measured noise.

The externally posted station correlations can be used for selection, or the
configured workflow can evaluate every common BHZ station.  All analysis
correlations, lags, noise estimates, residuals, and assessments are recalculated
from the SAC files.
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import math
import platform
import shutil
import sys
import time
import traceback
from dataclasses import dataclass, field
from datetime import datetime, timezone
from fractions import Fraction
from pathlib import Path
from typing import Any, Iterable, Sequence

import matplotlib

matplotlib.use("Agg")
import matplotlib.pyplot as plt
import numpy as np
import obspy
import openpyxl
import scipy
from obspy import UTCDateTime, read
from obspy.geodetics import gps2dist_azimuth, locations2degrees
from obspy.taup import TauPyModel
from scipy.signal import butter, detrend, resample_poly, sosfilt


ASSESS_SAME = "same_within_noise"
ASSESS_DIFFERENT = "different_beyond_noise"
ASSESS_INDETERMINATE = "indeterminate"
MISSING_SAC = -12345.0
STATION_MODE_SELECTED = "selected"
STATION_MODE_ALL = "all"
PLOTTED_PHASES = ("P", "pP", "sP", "PP", "PcP")


class AnalysisError(RuntimeError):
    """An expected, reportable scientific-input or processing failure."""


@dataclass(frozen=True)
class Event:
    event_id: int
    origin: UTCDateTime
    latitude: float
    longitude: float
    depth_km: float


@dataclass(frozen=True)
class Pair:
    label: str
    event1: Event
    event2: Event
    latitude: float
    longitude: float
    depth_km: float


@dataclass(frozen=True)
class Selection:
    pair_label: str
    event1: int
    event2: int
    network: str
    station: str
    posted_correlation: float | None
    source_figure: str
    selection_status: str
    notes: str = ""

    @property
    def station_id(self) -> str:
        return f"{self.network}.{self.station}" if self.network and self.station else ""


@dataclass
class ProcessedTrace:
    data: np.ndarray
    start_epoch: float
    sampling_hz: float
    native_sampling_hz: float
    path: Path
    network: str
    station: str
    location: str
    station_latitude: float
    station_longitude: float
    header_p_seconds_from_origin: float | None
    qc_flags: list[str] = field(default_factory=list)


@dataclass
class NoiseEstimate:
    chunks: list[np.ndarray]
    rms_values: np.ndarray
    median_rms: float
    p16_rms: float
    p84_rms: float
    usable_duration_seconds: float
    large_chunk_indices: list[int]


@dataclass
class StationAnalysis:
    result: dict[str, Any]
    noise_rows: list[dict[str, Any]]
    relative_plot_time: np.ndarray
    aligned_plot1: np.ndarray
    aligned_plot2: np.ndarray
    residual_time: np.ndarray
    residual: np.ndarray
    null_distribution: np.ndarray
    noise1: NoiseEstimate
    noise2: NoiseEstimate
    phase_plot_times: dict[str, tuple[float | None, float | None]] = field(
        default_factory=dict
    )


@dataclass(frozen=True)
class AnalysisRun:
    output_directory: Path
    elapsed_time_seconds: float
    cpu_time_seconds: float


def utc_now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def report_runtime(elapsed_time_seconds: float, cpu_time_seconds: float) -> None:
    print(f"Elapsed time: {elapsed_time_seconds:.3f} seconds")
    print(f"CPU time: {cpu_time_seconds:.3f} seconds")


def load_json(path: Path) -> dict[str, Any]:
    with path.open(encoding="utf-8") as handle:
        return json.load(handle)


def validate_config_paths(config: dict[str, Any]) -> None:
    for key in ("catalog_path", "supplementary_deck"):
        path = Path(config[key])
        if not path.is_file():
            raise AnalysisError(f"Configured {key} is not a file: {path}")
    for key in ("waveform_root", "figure_root"):
        path = Path(config[key])
        if not path.is_dir():
            raise AnalysisError(f"Configured {key} is not a directory: {path}")


def station_evaluation_mode(config: dict[str, Any]) -> str:
    mode = str(config.get("station_evaluation_mode", STATION_MODE_SELECTED)).strip().lower()
    if mode not in {STATION_MODE_SELECTED, STATION_MODE_ALL}:
        raise AnalysisError(
            "station_evaluation_mode must be 'selected' or 'all'"
        )
    return mode


def excluded_station_codes(config: dict[str, Any]) -> set[str]:
    """Return station codes excluded regardless of network prefix."""
    return {
        str(value).strip().upper().split(".")[-1]
        for value in config.get("excluded_stations", [])
        if str(value).strip()
    }


def station_is_excluded(station_id: str, excluded: set[str]) -> bool:
    return str(station_id).strip().upper().split(".")[-1] in excluded


def normalized_header_map(row: Sequence[Any]) -> dict[str, int]:
    return {
        str(value).strip().lower(): index
        for index, value in enumerate(row)
        if value is not None and str(value).strip()
    }


def _value(row: Sequence[Any], header: dict[str, int], name: str) -> Any:
    index = header.get(name.lower())
    return row[index] if index is not None and index < len(row) else None


def resolve_catalog(
    workbook_path: Path,
    pair_labels: Iterable[str],
    coordinate_tolerance_degrees: float,
    coordinate_tolerance_depth_km: float,
) -> dict[str, Pair]:
    """Resolve requested pairs and enforce pair/event coordinate consistency."""
    workbook = openpyxl.load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        event_sheet = workbook["events"]
        event_rows = event_sheet.iter_rows(values_only=True)
        event_header = normalized_header_map(next(event_rows))
        events: dict[int, Event] = {}
        for row in event_rows:
            raw_id = _value(row, event_header, "index")
            if raw_id is None:
                continue
            event_id = int(raw_id)
            events[event_id] = Event(
                event_id=event_id,
                origin=UTCDateTime(str(_value(row, event_header, "time"))),
                latitude=float(_value(row, event_header, "lat_best")),
                longitude=float(_value(row, event_header, "lon_best")),
                depth_km=float(_value(row, event_header, "depth_best")),
            )

        wanted = set(pair_labels)
        pair_sheet = workbook["pairs"]
        pair_rows = pair_sheet.iter_rows(values_only=True)
        pair_header = normalized_header_map(next(pair_rows))
        resolved: dict[str, Pair] = {}
        for row in pair_rows:
            label_raw = _value(row, pair_header, "label")
            if label_raw is None or str(label_raw).strip() not in wanted:
                continue
            label = str(label_raw).strip()
            event1_id = int(_value(row, pair_header, "index1"))
            event2_id = int(_value(row, pair_header, "index2"))
            if event1_id not in events or event2_id not in events:
                raise AnalysisError(f"{label}: workbook event row is missing")
            latitude = float(_value(row, pair_header, "lat"))
            longitude = float(_value(row, pair_header, "lon"))
            depth_km = float(_value(row, pair_header, "depth"))
            for event in (events[event1_id], events[event2_id]):
                if (
                    abs(event.latitude - latitude) > coordinate_tolerance_degrees
                    or abs(event.longitude - longitude) > coordinate_tolerance_degrees
                    or abs(event.depth_km - depth_km) > coordinate_tolerance_depth_km
                ):
                    raise AnalysisError(
                        f"{label}: pair coordinates do not match event {event.event_id}"
                    )
            resolved[label] = Pair(
                label,
                events[event1_id],
                events[event2_id],
                latitude,
                longitude,
                depth_km,
            )
        missing = wanted - resolved.keys()
        if missing:
            raise AnalysisError(f"Pair rows not found: {', '.join(sorted(missing))}")
        return resolved
    finally:
        workbook.close()


def load_selections(path: Path) -> list[Selection]:
    selections: list[Selection] = []
    with path.open(newline="", encoding="utf-8") as handle:
        for row_number, row in enumerate(csv.DictReader(handle), start=2):
            status = (row.get("selection_status") or "").strip().lower()
            correlation_text = (row.get("posted_correlation") or "").strip()
            correlation = float(correlation_text) if correlation_text else None
            try:
                event1 = int(row["event1"])
                event2 = int(row["event2"])
            except (KeyError, TypeError, ValueError) as exc:
                raise AnalysisError(
                    f"Invalid event IDs in selection row {row_number}"
                ) from exc
            selections.append(
                Selection(
                    pair_label=(row.get("pair_label") or "").strip(),
                    event1=event1,
                    event2=event2,
                    network=(row.get("network") or "").strip(),
                    station=(row.get("station") or "").strip(),
                    posted_correlation=correlation,
                    source_figure=(row.get("source_figure") or "").strip(),
                    selection_status=status,
                    notes=(row.get("notes") or "").strip(),
                )
            )
    return selections


def validate_selections(
    selections: Sequence[Selection], pairs: dict[str, Pair], threshold: float
) -> None:
    seen: set[tuple[str, str]] = set()
    for selection in selections:
        if selection.pair_label not in pairs:
            raise AnalysisError(f"Unknown selection pair {selection.pair_label}")
        pair = pairs[selection.pair_label]
        if (selection.event1, selection.event2) != (
            pair.event1.event_id,
            pair.event2.event_id,
        ):
            raise AnalysisError(
                f"{selection.pair_label}: selection event IDs disagree with catalog"
            )
        if selection.selection_status == "confirmed":
            if not selection.station_id or selection.posted_correlation is None:
                raise AnalysisError(
                    f"{selection.pair_label}: incomplete confirmed selection"
                )
            if selection.posted_correlation < threshold:
                raise AnalysisError(
                    f"{selection.pair_label} {selection.station_id}: posted correlation below threshold"
                )
            key = (selection.pair_label, selection.station_id)
            if key in seen:
                raise AnalysisError(f"Duplicate selection: {key[0]} {key[1]}")
            seen.add(key)
        elif selection.selection_status != "review":
            raise AnalysisError(
                f"Unrecognized selection status {selection.selection_status!r}"
            )


def validate_selection_figures(
    selections: Sequence[Selection], figure_root: Path
) -> None:
    missing = sorted(
        {
            selection.source_figure
            for selection in selections
            if selection.source_figure
            and not (figure_root / selection.source_figure).is_file()
        }
    )
    if missing:
        raise AnalysisError(f"Selection source figures not found: {', '.join(missing)}")


def event_directory(root: Path, event: Event) -> Path:
    key = event.origin.strftime("%Y%m%d_%H%M%S")
    direct = root / key
    if direct.is_dir():
        return direct
    matches = sorted(path for path in root.glob(f"{key}*") if path.is_dir())
    if len(matches) != 1:
        raise AnalysisError(
            f"Event {event.event_id}: expected one waveform directory matching {key}, found {len(matches)}"
        )
    return matches[0]


def _header_float(sac: Any, name: str) -> float | None:
    value = getattr(sac, name, None)
    try:
        parsed = float(value)
    except (TypeError, ValueError):
        return None
    if not math.isfinite(parsed) or abs(parsed - MISSING_SAC) < 1e-3:
        return None
    return parsed


def index_bhz_traces(event_dir: Path) -> dict[str, list[dict[str, Any]]]:
    indexed: dict[str, list[dict[str, Any]]] = {}
    for path in sorted(event_dir.rglob("*")):
        if not path.is_file() or path.suffix.lower() != ".sac":
            continue
        try:
            trace = read(str(path), headonly=True)[0]
        except Exception:
            continue
        if str(trace.stats.channel).upper() != "BHZ":
            continue
        station_id = f"{trace.stats.network}.{trace.stats.station}"
        sac = trace.stats.sac
        indexed.setdefault(station_id, []).append(
            {
                "path": path,
                "network": str(trace.stats.network),
                "station": str(trace.stats.station),
                "location": str(trace.stats.location or ""),
                "latitude": _header_float(sac, "stla"),
                "longitude": _header_float(sac, "stlo"),
                "npts": int(trace.stats.npts),
            }
        )
    return indexed


def choose_trace_pair(
    candidates1: Sequence[dict[str, Any]],
    candidates2: Sequence[dict[str, Any]],
    coordinate_tolerance_degrees: float,
) -> tuple[Path, Path]:
    """Match exact network.station and allow location changes only at equal coordinates."""
    matches: list[tuple[dict[str, Any], dict[str, Any]]] = []
    for first in candidates1:
        for second in candidates2:
            coordinates = (
                first["latitude"],
                first["longitude"],
                second["latitude"],
                second["longitude"],
            )
            if any(value is None for value in coordinates):
                continue
            if (
                abs(first["latitude"] - second["latitude"])
                <= coordinate_tolerance_degrees
                and abs(first["longitude"] - second["longitude"])
                <= coordinate_tolerance_degrees
            ):
                matches.append((first, second))
    if not matches:
        raise AnalysisError("No coordinate-consistent BHZ trace pair")
    matches.sort(
        key=lambda pair: (min(pair[0]["npts"], pair[1]["npts"]), str(pair[0]["path"])),
        reverse=True,
    )
    best = matches[0]
    best_score = min(best[0]["npts"], best[1]["npts"])
    tied = [
        pair for pair in matches if min(pair[0]["npts"], pair[1]["npts"]) == best_score
    ]
    if len(tied) > 1:
        exact_locations = [
            pair for pair in tied if pair[0]["location"] == pair[1]["location"]
        ]
        if len(exact_locations) == 1:
            best = exact_locations[0]
        elif len(tied) > 1:
            raise AnalysisError(
                "Ambiguous multiple coordinate-consistent BHZ trace pairs"
            )
    return best[0]["path"], best[1]["path"]


def all_station_candidates(
    pair: Pair,
    index1: dict[str, list[dict[str, Any]]],
    index2: dict[str, list[dict[str, Any]]],
    excluded: set[str] | None = None,
) -> list[Selection]:
    """Build deterministic candidates from the exact network.station intersection."""
    candidates: list[Selection] = []
    for station_id in sorted(set(index1).intersection(index2)):
        if excluded and station_is_excluded(station_id, excluded):
            continue
        network, station = station_id.split(".", 1)
        candidates.append(
            Selection(
                pair_label=pair.label,
                event1=pair.event1.event_id,
                event2=pair.event2.event_id,
                network=network,
                station=station,
                posted_correlation=None,
                source_figure="",
                selection_status=STATION_MODE_ALL,
            )
        )
    return candidates


def rows_meeting_correlation_threshold(
    rows: Sequence[dict[str, Any]], threshold: float
) -> list[dict[str, Any]]:
    return [
        {"selection_correlation_threshold": threshold, **row}
        for row in rows
        if float(row["new_correlation"]) >= threshold
    ]


def stations_meeting_correlation_threshold(
    stations: Sequence[StationAnalysis], threshold: float
) -> list[StationAnalysis]:
    """Return analyzed stations whose newly calculated CC meets the threshold."""
    return [
        station
        for station in stations
        if float(station.result["new_correlation"]) >= threshold
    ]


def report_station_progress(
    completed: int,
    total: int,
    pair_label: str = "",
    station_id: str = "",
    status: str = "",
    interval: int = 20,
) -> None:
    if completed <= 0 or (completed % interval != 0 and completed != total):
        return
    details = (
        f" | {pair_label} {station_id} | {status}"
        if pair_label and station_id and status
        else ""
    )
    print(
        f"Station evaluations completed: {completed}/{total}{details}",
        flush=True,
    )


def requested_phase_arrivals(
    model: TauPyModel,
    event: Event,
    station_latitude: float,
    station_longitude: float,
    phases: Sequence[str] = PLOTTED_PHASES,
) -> tuple[dict[str, float], float]:
    """Return the earliest named TauP arrival for each requested phase."""
    distance_degrees = locations2degrees(
        event.latitude, event.longitude, station_latitude, station_longitude
    )
    arrivals = model.get_travel_times(
        source_depth_in_km=event.depth_km,
        distance_in_degree=distance_degrees,
        phase_list=list(phases),
    )
    earliest: dict[str, float] = {}
    for phase in phases:
        matching = [arrival for arrival in arrivals if arrival.name == phase]
        if matching:
            earliest[phase] = float(event.origin) + min(
                float(arrival.time) for arrival in matching
            )
    if "P" in phases and "P" not in earliest:
        raise AnalysisError("TauP ak135 has no direct P arrival")
    return earliest, float(distance_degrees)


def direct_p_arrival(
    model: TauPyModel, event: Event, station_latitude: float, station_longitude: float
) -> tuple[float, float]:
    arrivals, distance_degrees = requested_phase_arrivals(
        model, event, station_latitude, station_longitude, ("P",)
    )
    return arrivals["P"], distance_degrees


def aligned_phase_plot_times(
    arrivals1: dict[str, float],
    arrivals2: dict[str, float],
    direct_p1: float,
    direct_p2: float,
    lag_seconds_y_t_plus_lag: float,
) -> dict[str, tuple[float | None, float | None]]:
    """Map arrivals onto the displayed axes, including the event-2 lag shift."""
    return {
        phase: (
            arrivals1[phase] - direct_p1 if phase in arrivals1 else None,
            arrivals2[phase] - direct_p2 - lag_seconds_y_t_plus_lag
            if phase in arrivals2
            else None,
        )
        for phase in PLOTTED_PHASES
        if phase in arrivals1 or phase in arrivals2
    }


def paired_epicentral_distance_degrees(
    pair: Pair,
    station1_latitude: float,
    station1_longitude: float,
    station2_latitude: float,
    station2_longitude: float,
) -> float:
    """Return the mean event-to-station angular distance for a matched trace pair."""
    distance1 = locations2degrees(
        pair.event1.latitude,
        pair.event1.longitude,
        station1_latitude,
        station1_longitude,
    )
    distance2 = locations2degrees(
        pair.event2.latitude,
        pair.event2.longitude,
        station2_latitude,
        station2_longitude,
    )
    return 0.5 * (float(distance1) + float(distance2))


def _clipping_detected(data: np.ndarray, repeat_count: int) -> bool:
    if len(data) == 0 or repeat_count <= 1:
        return False
    extrema = (data == np.nanmax(data)) | (data == np.nanmin(data))
    run = 0
    for value in extrema:
        run = run + 1 if value else 0
        if run >= repeat_count:
            return True
    return False


def preprocess_trace(path: Path, config: dict[str, Any]) -> ProcessedTrace:
    stream = read(str(path))
    if len(stream) != 1:
        raise AnalysisError(f"Expected one trace in {path}, found {len(stream)}")
    trace = stream[0]
    if str(trace.stats.channel).upper() != "BHZ":
        raise AnalysisError(f"Not a BHZ trace: {path}")
    raw = np.ma.asarray(trace.data)
    if np.ma.isMaskedArray(raw) and np.any(np.ma.getmaskarray(raw)):
        raise AnalysisError(f"Trace contains gaps: {path}")
    data = np.asarray(raw, dtype=np.float64)
    if len(data) < 2 or not np.all(np.isfinite(data)):
        raise AnalysisError(f"Trace has insufficient or nonfinite samples: {path}")
    native_hz = float(trace.stats.sampling_rate)
    low, high = map(float, config["bandpass_hz"])
    if not math.isfinite(native_hz) or native_hz <= 2.0 * high:
        raise AnalysisError(f"Sampling rate {native_hz:g} cannot support {high:g} Hz")
    qc_flags: list[str] = []
    if _clipping_detected(data, int(config["clipping_repeat_count"])):
        qc_flags.append("clipping_indicator")
    data = detrend(data, type="linear")
    taper_samples = min(
        len(data), int(round(float(config["start_taper_seconds"]) * native_hz))
    )
    if taper_samples > 1:
        phase = np.linspace(0.0, np.pi, taper_samples)
        data[:taper_samples] *= 0.5 * (1.0 - np.cos(phase))
    sos = butter(
        int(config["filter_order"]),
        [low, high],
        btype="bandpass",
        fs=native_hz,
        output="sos",
    )
    filtered = sosfilt(sos, data)
    target_hz = float(config["target_sampling_hz"])
    ratio = Fraction(target_hz / native_hz).limit_denominator(10000)
    resampled = resample_poly(filtered, ratio.numerator, ratio.denominator)
    sac = trace.stats.sac
    station_latitude = _header_float(sac, "stla")
    station_longitude = _header_float(sac, "stlo")
    if station_latitude is None or station_longitude is None:
        raise AnalysisError(f"Missing SAC station coordinates: {path}")
    header_p_seconds_from_origin: float | None = None
    t0 = _header_float(sac, "t0")
    kt0 = str(getattr(sac, "kt0", "")).strip().upper()
    if t0 is not None and kt0.startswith("P"):
        # In this data set t0 stores P travel time from the catalog origin,
        # while the record begins about 100 s before that origin.  Preserve the
        # documented data-set convention instead of treating t0 as time after
        # the SAC record reference.
        header_p_seconds_from_origin = t0
    return ProcessedTrace(
        data=np.asarray(resampled, dtype=np.float64),
        start_epoch=float(trace.stats.starttime),
        sampling_hz=target_hz,
        native_sampling_hz=native_hz,
        path=path,
        network=str(trace.stats.network),
        station=str(trace.stats.station),
        location=str(trace.stats.location or ""),
        station_latitude=station_latitude,
        station_longitude=station_longitude,
        header_p_seconds_from_origin=header_p_seconds_from_origin,
        qc_flags=qc_flags,
    )


def window_times(window: Sequence[float], sampling_hz: float) -> np.ndarray:
    start, end = map(float, window)
    count_float = (end - start) * sampling_hz
    count = int(round(count_float))
    if count <= 0 or not math.isclose(count_float, count, abs_tol=1e-7):
        raise AnalysisError(
            f"Window {window} is not exact on the {sampling_hz:g} Hz grid"
        )
    return start + np.arange(count, dtype=np.float64) / sampling_hz


def extract_at_epochs(trace: ProcessedTrace, epochs: np.ndarray) -> np.ndarray:
    positions = (np.asarray(epochs) - trace.start_epoch) * trace.sampling_hz
    if positions.size and (
        positions[0] < -1e-7 or positions[-1] > len(trace.data) - 1 + 1e-7
    ):
        raise AnalysisError(f"Requested window is outside {trace.path.name}")
    positions = np.clip(positions, 0.0, len(trace.data) - 1.0)
    base = np.arange(len(trace.data), dtype=np.float64)
    return np.interp(positions, base, trace.data)


def extract_relative(
    trace: ProcessedTrace, arrival_epoch: float, relative_times: np.ndarray
) -> np.ndarray:
    return extract_at_epochs(trace, arrival_epoch + np.asarray(relative_times))


def pearson_signed(first: np.ndarray, second: np.ndarray) -> float:
    first_centered = np.asarray(first, dtype=float) - np.mean(first)
    second_centered = np.asarray(second, dtype=float) - np.mean(second)
    denominator = np.linalg.norm(first_centered) * np.linalg.norm(second_centered)
    if denominator == 0.0:
        return math.nan
    return float(np.dot(first_centered, second_centered) / denominator)


def signed_lag_correlation(
    first: ProcessedTrace,
    second: ProcessedTrace,
    arrival1_epoch: float,
    arrival2_epoch: float,
    window: Sequence[float],
    max_lag_seconds: float,
) -> tuple[float, float, bool, np.ndarray, np.ndarray]:
    """Maximize signed CC for x(t) versus y(t + lag), then refine parabolically."""
    relative = window_times(window, first.sampling_hz)
    first_window = extract_relative(first, arrival1_epoch, relative)
    max_samples = int(round(max_lag_seconds * first.sampling_hz))
    lag_samples = np.arange(-max_samples, max_samples + 1)
    correlations = np.empty(len(lag_samples), dtype=float)
    for index, lag_sample in enumerate(lag_samples):
        second_window = extract_relative(
            second,
            arrival2_epoch,
            relative + lag_sample / first.sampling_hz,
        )
        correlations[index] = pearson_signed(first_window, second_window)
    if not np.any(np.isfinite(correlations)):
        raise AnalysisError("Correlation is undefined at every searched lag")
    best_index = int(np.nanargmax(correlations))
    boundary = best_index in (0, len(correlations) - 1)
    refined_sample = float(lag_samples[best_index])
    if not boundary:
        left, center, right = correlations[best_index - 1 : best_index + 2]
        denominator = left - 2.0 * center + right
        if math.isfinite(denominator) and abs(denominator) > np.finfo(float).eps:
            offset = 0.5 * (left - right) / denominator
            if abs(offset) <= 1.0:
                refined_sample += float(offset)
    lag_seconds = refined_sample / first.sampling_hz
    aligned_second = extract_relative(second, arrival2_epoch, relative + lag_seconds)
    correlation = pearson_signed(first_window, aligned_second)
    return lag_seconds, correlation, boundary, first_window, aligned_second


def symmetric_residual(
    first: np.ndarray, second: np.ndarray
) -> tuple[np.ndarray, float, float, float]:
    scale1 = float(np.sqrt(np.mean(np.square(first))))
    scale2 = float(np.sqrt(np.mean(np.square(second))))
    if scale1 <= 0.0 or scale2 <= 0.0:
        raise AnalysisError("A correlation-window RMS normalization is zero")
    residual = first / scale1 - second / scale2
    return residual, scale1, scale2, float(np.sqrt(np.mean(np.square(residual))))


def chunk_noise(
    trace: ProcessedTrace,
    arrival_epoch: float,
    config: dict[str, Any],
) -> NoiseEstimate:
    chunk_seconds = float(config["noise_chunk_seconds"])
    pre_p_guard_seconds = float(config["noise_pre_p_guard_seconds"])
    if pre_p_guard_seconds < 0.0:
        raise AnalysisError("noise_pre_p_guard_seconds must be nonnegative")
    sample_count = int(round(chunk_seconds * trace.sampling_hz))
    usable_start = trace.start_epoch + float(config["filter_startup_exclusion_seconds"])
    usable_end = arrival_epoch - pre_p_guard_seconds
    usable_duration = max(0.0, usable_end - usable_start)
    count = int(math.floor(usable_duration / chunk_seconds + 1e-10))
    chunks: list[np.ndarray] = []
    rms_values: list[float] = []
    for index in range(count):
        # Anchor chunks backward from the guard boundary so the final complete
        # noise window ends exactly pre_p_guard_seconds before predicted P.
        start = usable_end - (count - index) * chunk_seconds
        epochs = start + np.arange(sample_count, dtype=float) / trace.sampling_hz
        try:
            chunk = extract_at_epochs(trace, epochs)
        except AnalysisError:
            continue
        if not np.all(np.isfinite(chunk)) or _clipping_detected(
            chunk, int(config["clipping_repeat_count"])
        ):
            continue
        chunks.append(chunk)
        rms_values.append(float(np.sqrt(np.mean(np.square(chunk)))))
    values = np.asarray(rms_values, dtype=float)
    if len(values):
        median = float(np.median(values))
        p16, p84 = map(float, np.percentile(values, [16.0, 84.0]))
        mad = float(np.median(np.abs(values - median)))
        cutoff = median + float(config["large_noise_chunk_mad_multiplier"]) * mad
        large = np.flatnonzero(values > cutoff).astype(int).tolist() if mad > 0 else []
    else:
        median = p16 = p84 = math.nan
        large = []
    return NoiseEstimate(chunks, values, median, p16, p84, usable_duration, large)


def empirical_noise_null(
    chunks1: Sequence[np.ndarray],
    chunks2: Sequence[np.ndarray],
    scale1: float,
    scale2: float,
) -> np.ndarray:
    if not chunks1 or not chunks2:
        return np.empty(0, dtype=float)
    first = np.asarray(chunks1, dtype=float) / scale1
    second = np.asarray(chunks2, dtype=float) / scale2
    if first.ndim != 2 or second.ndim != 2 or first.shape[1] != second.shape[1]:
        raise AnalysisError(
            "Noise chunks do not share one exact duration and sampling grid"
        )
    sample_count = first.shape[1]
    first_energy = np.mean(first * first, axis=1)[:, None]
    second_energy = np.mean(second * second, axis=1)[None, :]
    cross = first @ second.T / sample_count
    squared = np.maximum(0.0, first_energy + second_energy - 2.0 * cross)
    return np.sqrt(squared).ravel()


def classify_trace(
    observed_residual_rms: float,
    null_distribution: np.ndarray,
    valid_chunk_count1: int,
    valid_chunk_count2: int,
    minimum_chunks: int,
    material_qc_flags: Sequence[str],
) -> tuple[str, float]:
    if len(null_distribution):
        threshold = float(np.percentile(null_distribution, 95.0))
    else:
        threshold = math.nan
    if (
        valid_chunk_count1 < minimum_chunks
        or valid_chunk_count2 < minimum_chunks
        or material_qc_flags
        or not math.isfinite(threshold)
    ):
        return ASSESS_INDETERMINATE, threshold
    assessment = ASSESS_SAME if observed_residual_rms <= threshold else ASSESS_DIFFERENT
    return assessment, threshold


def pair_bootstrap(
    stations: Sequence[StationAnalysis],
    iterations: int,
    seed: int,
    minimum_stations: int,
) -> dict[str, Any]:
    usable = [
        station
        for station in stations
        if station.result["trace_assessment"] != ASSESS_INDETERMINATE
        and len(station.null_distribution)
    ]
    if len(usable) < minimum_stations:
        return {
            "classifiable_station_count": len(usable),
            "station_median_residual_rms": math.nan,
            "station_median_residual_ci16": math.nan,
            "station_median_residual_ci84": math.nan,
            "pair_null_p95": math.nan,
            "pair_assessment": ASSESS_INDETERMINATE,
        }
    rng = np.random.default_rng(seed)
    observed = np.asarray(
        [station.result["residual_rms"] for station in usable], dtype=float
    )
    observed_median = float(np.median(observed))
    observed_bootstrap = np.empty(iterations, dtype=float)
    null_bootstrap = np.empty(iterations, dtype=float)
    station_count = len(usable)
    for index in range(iterations):
        chosen = rng.integers(0, station_count, size=station_count)
        observed_bootstrap[index] = np.median(observed[chosen])
        null_values = [
            usable[station_index].null_distribution[
                rng.integers(0, len(usable[station_index].null_distribution))
            ]
            for station_index in chosen
        ]
        null_bootstrap[index] = np.median(null_values)
    ci16, ci84 = np.percentile(observed_bootstrap, [16.0, 84.0])
    null_p95 = float(np.percentile(null_bootstrap, 95.0))
    return {
        "classifiable_station_count": station_count,
        "station_median_residual_rms": observed_median,
        "station_median_residual_ci16": float(ci16),
        "station_median_residual_ci84": float(ci84),
        "pair_null_p95": null_p95,
        "pair_assessment": (
            ASSESS_SAME if observed_median <= null_p95 else ASSESS_DIFFERENT
        ),
    }


def analyze_station(
    selection: Selection,
    pair: Pair,
    path1: Path,
    path2: Path,
    model: TauPyModel,
    config: dict[str, Any],
) -> StationAnalysis:
    trace1 = preprocess_trace(path1, config)
    trace2 = preprocess_trace(path2, config)
    station_tolerance = float(config["station_coordinate_tolerance_degrees"])
    if (
        abs(trace1.station_latitude - trace2.station_latitude) > station_tolerance
        or abs(trace1.station_longitude - trace2.station_longitude) > station_tolerance
    ):
        raise AnalysisError("Station coordinates differ between event traces")
    arrivals1, distance1 = requested_phase_arrivals(
        model, pair.event1, trace1.station_latitude, trace1.station_longitude
    )
    arrivals2, distance2 = requested_phase_arrivals(
        model, pair.event2, trace2.station_latitude, trace2.station_longitude
    )
    arrival1 = arrivals1["P"]
    arrival2 = arrivals2["P"]
    qc_flags = list(dict.fromkeys(trace1.qc_flags + trace2.qc_flags))
    header_delta1 = (
        arrival1 - float(pair.event1.origin) - trace1.header_p_seconds_from_origin
        if trace1.header_p_seconds_from_origin is not None
        else math.nan
    )
    header_delta2 = (
        arrival2 - float(pair.event2.origin) - trace2.header_p_seconds_from_origin
        if trace2.header_p_seconds_from_origin is not None
        else math.nan
    )
    tolerance = float(config["header_p_tolerance_seconds"])
    if math.isfinite(header_delta1) and abs(header_delta1) > tolerance:
        qc_flags.append("event1_header_p_discrepancy")
    if math.isfinite(header_delta2) and abs(header_delta2) > tolerance:
        qc_flags.append("event2_header_p_discrepancy")

    lag, correlation, lag_boundary, correlation1, correlation2 = signed_lag_correlation(
        trace1,
        trace2,
        arrival1,
        arrival2,
        config["correlation_window_seconds"],
        float(config["lag_search_seconds"]),
    )
    if lag_boundary:
        qc_flags.append("lag_at_search_boundary")
    phase_plot_times = aligned_phase_plot_times(
        arrivals1, arrivals2, arrival1, arrival2, lag
    )
    # The correlation-window mean removal applies both to CC and RMS normalization.
    correlation1 = correlation1 - np.mean(correlation1)
    correlation2 = correlation2 - np.mean(correlation2)
    _, scale1, scale2, _ = symmetric_residual(correlation1, correlation2)

    residual_time = window_times(config["residual_window_seconds"], trace1.sampling_hz)
    signal1 = extract_relative(trace1, arrival1, residual_time)
    signal2 = extract_relative(trace2, arrival2, residual_time + lag)
    residual = signal1 / scale1 - signal2 / scale2
    residual_rms = float(np.sqrt(np.mean(np.square(residual))))

    noise1 = chunk_noise(trace1, arrival1, config)
    noise2 = chunk_noise(trace2, arrival2, config)
    null_distribution = empirical_noise_null(
        noise1.chunks, noise2.chunks, scale1, scale2
    )
    material_flags = [
        flag
        for flag in qc_flags
        if flag in {"clipping_indicator", "lag_at_search_boundary"}
    ]
    assessment, null_p95 = classify_trace(
        residual_rms,
        null_distribution,
        len(noise1.chunks),
        len(noise2.chunks),
        int(config["minimum_noise_chunks"]),
        material_flags,
    )
    normalized_noise1 = noise1.median_rms / scale1
    normalized_noise2 = noise2.median_rms / scale2
    denominator = math.sqrt(normalized_noise1**2 + normalized_noise2**2)
    r_sym = residual_rms / denominator if denominator > 0.0 else math.nan
    if (
        selection.posted_correlation is not None
        and correlation + 0.005 < selection.posted_correlation
    ):
        qc_flags.append("new_correlation_below_posted")

    plot_time = window_times(config["plot_window_seconds"], trace1.sampling_hz)
    plot1 = extract_relative(trace1, arrival1, plot_time) / scale1
    plot2 = extract_relative(trace2, arrival2, plot_time + lag) / scale2
    distance_m, azimuth_degrees, _ = gps2dist_azimuth(
        pair.latitude, pair.longitude, trace1.station_latitude, trace1.station_longitude
    )
    result: dict[str, Any] = {
        "pair_label": pair.label,
        "event1": pair.event1.event_id,
        "event2": pair.event2.event_id,
        "network": selection.network,
        "station": selection.station,
        "station_id": selection.station_id,
        "posted_correlation": selection.posted_correlation,
        "new_correlation": correlation,
        "lag_seconds_y_t_plus_lag": lag,
        "lag_at_boundary": lag_boundary,
        "trace1_path": str(path1),
        "trace2_path": str(path2),
        "trace1_location": trace1.location,
        "trace2_location": trace2.location,
        "station_latitude": trace1.station_latitude,
        "station_longitude": trace1.station_longitude,
        "epicentral_distance_degrees": 0.5 * (distance1 + distance2),
        "epicentral_distance_km": distance_m / 1000.0,
        "azimuth_degrees": azimuth_degrees,
        "calculated_p1_epoch": arrival1,
        "calculated_p2_epoch": arrival2,
        "calculated_p1_iso": str(UTCDateTime(arrival1)),
        "calculated_p2_iso": str(UTCDateTime(arrival2)),
        "calculated_minus_header_p1_seconds": header_delta1,
        "calculated_minus_header_p2_seconds": header_delta2,
        "native_sampling_hz1": trace1.native_sampling_hz,
        "native_sampling_hz2": trace2.native_sampling_hz,
        "analysis_sampling_hz": trace1.sampling_hz,
        "scale_s1_stored_amplitude_units": scale1,
        "scale_s2_stored_amplitude_units": scale2,
        "scale_ratio_s1_over_s2": scale1 / scale2,
        "noise_median1_stored_amplitude_units": noise1.median_rms,
        "noise_p16_1_stored_amplitude_units": noise1.p16_rms,
        "noise_p84_1_stored_amplitude_units": noise1.p84_rms,
        "noise_median2_stored_amplitude_units": noise2.median_rms,
        "noise_p16_2_stored_amplitude_units": noise2.p16_rms,
        "noise_p84_2_stored_amplitude_units": noise2.p84_rms,
        "normalized_noise1": normalized_noise1,
        "normalized_noise2": normalized_noise2,
        "snr1_scale_over_noise": scale1 / noise1.median_rms
        if noise1.median_rms > 0
        else math.nan,
        "snr2_scale_over_noise": scale2 / noise2.median_rms
        if noise2.median_rms > 0
        else math.nan,
        "usable_noise_duration1_seconds": noise1.usable_duration_seconds,
        "usable_noise_duration2_seconds": noise2.usable_duration_seconds,
        "valid_noise_chunks1": len(noise1.chunks),
        "valid_noise_chunks2": len(noise2.chunks),
        "large_noise_chunks1": len(noise1.large_chunk_indices),
        "large_noise_chunks2": len(noise2.large_chunk_indices),
        "residual_rms": residual_rms,
        "r_sym": r_sym,
        "empirical_null_count": len(null_distribution),
        "empirical_null_p95": null_p95,
        "trace_assessment": assessment,
        "qc_flags": ";".join(dict.fromkeys(qc_flags)),
    }
    noise_rows: list[dict[str, Any]] = []
    for event_number, estimate, scale in ((1, noise1, scale1), (2, noise2, scale2)):
        for index, rms in enumerate(estimate.rms_values):
            noise_rows.append(
                {
                    "pair_label": pair.label,
                    "station_id": selection.station_id,
                    "event_number": event_number,
                    "event_id": pair.event1.event_id
                    if event_number == 1
                    else pair.event2.event_id,
                    "chunk_index": index,
                    "rms_stored_amplitude_units": rms,
                    "normalized_rms": rms / scale,
                    "unusually_large": index in estimate.large_chunk_indices,
                }
            )
    return StationAnalysis(
        result,
        noise_rows,
        plot_time,
        plot1,
        plot2,
        residual_time,
        residual,
        null_distribution,
        noise1,
        noise2,
        phase_plot_times,
    )


def station_trace_label(result: dict[str, Any]) -> str:
    """Return the compact annotation used on each plotted station trace."""
    return (
        f"{result['station_id']}  dist={result['epicentral_distance_degrees']:.1f}\N{DEGREE SIGN}  "
        f"az={result['azimuth_degrees']:.1f}\N{DEGREE SIGN}\n"
        f"CC={result['new_correlation']:.2f}  "
        f"shift={result['lag_seconds_y_t_plus_lag']:+.2f} s"
    )


def plot_time_axis_limits(relative_plot_time: np.ndarray) -> tuple[float, float]:
    """Return display limits implied by the half-open configured plot window."""
    if len(relative_plot_time) == 0:
        raise AnalysisError("Cannot plot an empty time window")
    x_min = float(relative_plot_time[0])
    if len(relative_plot_time) == 1:
        return x_min, x_min
    sample_interval = float(np.median(np.diff(relative_plot_time)))
    return x_min, float(relative_plot_time[-1]) + sample_interval


def mark_phase_arrivals_on_overlay(
    axis: Any, station: StationAnalysis, baseline: float | None = None
) -> None:
    """Draw labeled event-1 arrival markers on an overlay axis."""
    x_min = float(station.relative_plot_time[0])
    x_max = float(station.relative_plot_time[-1])
    for phase, event_times in station.phase_plot_times.items():
        arrival_time = event_times[0]
        if arrival_time is None or not x_min <= arrival_time <= x_max:
            continue
        if baseline is None:
            axis.axvline(arrival_time, color="tab:blue", linewidth=0.8, alpha=0.65)
            axis.text(
                arrival_time,
                0.98,
                phase,
                color="tab:blue",
                fontsize=7,
                rotation=90,
                ha="right",
                va="top",
                transform=axis.get_xaxis_transform(),
            )
        else:
            axis.vlines(
                arrival_time,
                baseline - 2.35,
                baseline + 2.35,
                color="tab:blue",
                linewidth=0.7,
                alpha=0.6,
            )
            axis.text(
                arrival_time,
                baseline + 2.4,
                phase,
                color="tab:blue",
                fontsize=5,
                rotation=90,
                ha="center",
                va="bottom",
            )


def plot_pair_overlay(
    pair: Pair,
    stations: Sequence[StationAnalysis],
    output_path: Path,
    correlation_threshold: float,
) -> None:
    ordered = sorted(
        stations,
        key=lambda station: station.result["epicentral_distance_degrees"],
        reverse=True,
    )
    height = max(5.0, 1.0 + 0.65 * len(ordered))
    figure, axis = plt.subplots(figsize=(13, height), constrained_layout=True)
    for row, station in enumerate(ordered):
        baseline = (len(ordered) - 1 - row) * 6.0
        x_min, x_max = plot_time_axis_limits(station.relative_plot_time)
        axis.plot(
            station.relative_plot_time,
            station.aligned_plot1 + baseline,
            color="tab:blue",
            linewidth=0.8,
        )
        axis.plot(
            station.relative_plot_time,
            station.aligned_plot2 + baseline,
            color="tab:red",
            linewidth=0.8,
        )
        mark_phase_arrivals_on_overlay(axis, station, baseline)
        axis.text(
            x_min + 0.008 * (x_max - x_min),
            baseline + 2.55,
            station_trace_label(station.result),
            ha="left",
            va="top",
            fontsize=8,
        )
    axis.axvspan(-1.0, 9.0, color="0.8", alpha=0.25)
    if ordered:
        axis.set_xlim(*plot_time_axis_limits(ordered[0].relative_plot_time))
    axis.set_yticks([])
    axis.set_xlabel("Time relative to calculated P (s)")
    axis.set_title(
        f"{pair.label}: event {pair.event1.event_id} (blue) vs {pair.event2.event_id} (red)\n"
        "Aligned and symmetrically normalized; uniform 6-unit baselines; "
        f"CC >= {correlation_threshold:g}"
    )
    figure.savefig(output_path, dpi=180, bbox_inches="tight")
    plt.close(figure)


def plot_station_diagnostic(station: StationAnalysis, output_path: Path) -> None:
    result = station.result
    figure, (overlay, residual_axis) = plt.subplots(
        2,
        1,
        figsize=(13, 7),
        constrained_layout=True,
        height_ratios=(2, 1),
    )
    overlay.plot(
        station.relative_plot_time,
        station.aligned_plot1,
        color="tab:blue",
        label="event 1",
    )
    overlay.plot(
        station.relative_plot_time,
        station.aligned_plot2,
        color="tab:red",
        label="event 2",
    )
    overlay.axvspan(-1.0, 9.0, color="0.7", alpha=0.25, label="CC window")
    overlay.axvspan(9.0, 29.0, color="0.85", alpha=0.2, label="residual only")
    mark_phase_arrivals_on_overlay(overlay, station)
    overlay.set_xlim(*plot_time_axis_limits(station.relative_plot_time))
    overlay.set_title("Aligned normalized traces")
    overlay.text(
        0.015,
        0.975,
        station_trace_label(result),
        transform=overlay.transAxes,
        ha="left",
        va="top",
        fontsize=9,
    )
    overlay.legend(fontsize=8, ncol=2, loc="upper right")

    residual_axis.plot(
        station.residual_time, station.residual, color="0.2", linewidth=0.8
    )
    residual_axis.axhline(0.0, color="0.6", linewidth=0.6)
    residual_axis.set_title("Normalized residual, [P-1, P+29)")
    residual_axis.set_xlabel("Time relative to P (s)")
    figure.suptitle(
        f"{result['pair_label']}: event {result['event1']} (blue) vs "
        f"{result['event2']} (red)",
        fontsize=11,
    )
    figure.savefig(output_path, dpi=180)
    plt.close(figure)


def plot_pair_summary(pair_result: dict[str, Any], output_path: Path) -> None:
    figure, axis = plt.subplots(figsize=(8, 4.5), constrained_layout=True)
    axis.axis("off")
    rows = [
        ("Pair", pair_result["pair_label"]),
        (
            "Stations analyzed / candidates",
            f"{pair_result['station_count']} / {pair_result['candidate_station_count']}",
        ),
        (
            "Stations meeting CC threshold",
            pair_result["correlation_threshold_station_count"],
        ),
        ("Classifiable stations", pair_result["classifiable_station_count"]),
        ("Median new correlation", f"{pair_result['median_new_correlation']:.3f}"),
        (
            "Exceptions / QC-flagged",
            f"{pair_result['exception_count']} / {pair_result['qc_flagged_station_count']}",
        ),
    ]
    table = axis.table(
        cellText=rows, colLabels=["Metric", "Value"], loc="center", cellLoc="left"
    )
    table.auto_set_font_size(False)
    table.set_fontsize(11)
    table.scale(1.0, 1.6)
    axis.set_title(
        f"{pair_result['pair_label']}: events {pair_result['event1']} and {pair_result['event2']}",
        fontsize=14,
    )
    figure.savefig(output_path, dpi=180)
    plt.close(figure)


def csv_safe(value: Any) -> Any:
    if isinstance(value, (np.floating, float)) and not math.isfinite(float(value)):
        return ""
    if isinstance(value, (np.integer,)):
        return int(value)
    return value


def write_csv_rows(
    path: Path, rows: Sequence[dict[str, Any]], fieldnames: Sequence[str] | None = None
) -> None:
    if fieldnames is None:
        fieldnames = list(dict.fromkeys(key for row in rows for key in row))
    with path.open("x", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        for row in rows:
            writer.writerow({key: csv_safe(row.get(key, "")) for key in fieldnames})


def write_strict_json(path: Path, value: Any) -> None:
    """Validate strict JSON completely before exclusively creating the file."""
    serialized = json.dumps(value, indent=2, allow_nan=False) + "\n"
    with path.open("x", encoding="utf-8") as handle:
        handle.write(serialized)


def create_timestamped_output(root: Path, now: datetime | None = None) -> Path:
    root.mkdir(parents=True, exist_ok=True)
    timestamp = (now or datetime.now()).strftime("%Y%m%d_%H%M%S")
    destination = root / timestamp
    destination.mkdir(parents=False, exist_ok=False)
    for name in ("pair_plots", "station_diagnostics", "pair_summaries"):
        (destination / name).mkdir()
    return destination


def file_fingerprint(path: Path) -> dict[str, Any]:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for block in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(block)
    stat = path.stat()
    return {
        "path": str(path),
        "size_bytes": stat.st_size,
        "mtime_ns": stat.st_mtime_ns,
        "sha256": digest.hexdigest(),
    }


def software_versions() -> dict[str, str]:
    return {
        "python": platform.python_version(),
        "numpy": np.__version__,
        "scipy": scipy.__version__,
        "obspy": obspy.__version__,
        "openpyxl": openpyxl.__version__,
        "matplotlib": matplotlib.__version__,
    }


def run_analysis(config_path: Path, selection_path: Path) -> AnalysisRun:
    elapsed_started = time.perf_counter()
    cpu_started = time.process_time()
    started = utc_now_iso()
    config = load_json(config_path)
    validate_config_paths(config)
    evaluation_mode = station_evaluation_mode(config)
    excluded = excluded_station_codes(config)
    correlation_threshold = float(config["selection_correlation_threshold"])
    pair_labels = [str(label) for label in config["pairs"]]
    pairs = resolve_catalog(
        Path(config["catalog_path"]),
        pair_labels,
        float(config["coordinate_tolerance_degrees"]),
        float(config["coordinate_tolerance_depth_km"]),
    )
    selections = [
        selection
        for selection in load_selections(selection_path)
        if not station_is_excluded(selection.station_id, excluded)
    ]
    validate_selections(selections, pairs, correlation_threshold)
    if evaluation_mode == STATION_MODE_SELECTED:
        validate_selection_figures(selections, Path(config["figure_root"]))
    output = create_timestamped_output(Path(config["output_root"]))
    shutil.copy2(config_path, output / "analysis_config.json")
    shutil.copy2(selection_path, output / "station_selection.csv")

    exceptions: list[dict[str, Any]] = []
    station_analyses: dict[str, list[StationAnalysis]] = {
        label: [] for label in pair_labels
    }
    noise_rows: list[dict[str, Any]] = []
    used_paths: set[Path] = {
        Path(config["catalog_path"]),
        selection_path,
        config_path,
        Path(config["supplementary_deck"]),
    }
    figure_root = Path(config["figure_root"])
    if evaluation_mode == STATION_MODE_SELECTED:
        for selection in selections:
            if selection.source_figure:
                used_paths.add(figure_root / selection.source_figure)
            if selection.selection_status == "review":
                exceptions.append(
                    {
                        "pair_label": selection.pair_label,
                        "station_id": selection.station_id,
                        "stage": "station_selection",
                        "exception": "selection_requires_review",
                        "details": selection.notes,
                    }
                )

    # Snapshot all fixed non-waveform inputs before reading or analyzing SAC data.
    fixed_fingerprints_before = {
        str(path): file_fingerprint(path)
        for path in sorted(used_paths)
        if path.is_file()
    }

    model = TauPyModel(model=str(config["taup_model"]))
    waveform_root = Path(config["waveform_root"])
    candidate_counts: dict[str, int] = {}
    pair_inputs: dict[
        str,
        tuple[
            dict[str, list[dict[str, Any]]],
            dict[str, list[dict[str, Any]]],
            list[Selection],
        ],
    ] = {}
    for pair_label in pair_labels:
        pair = pairs[pair_label]
        try:
            index1 = index_bhz_traces(event_directory(waveform_root, pair.event1))
            index2 = index_bhz_traces(event_directory(waveform_root, pair.event2))
        except Exception as exc:
            exceptions.append(
                {
                    "pair_label": pair_label,
                    "station_id": "",
                    "stage": "waveform_index",
                    "exception": type(exc).__name__,
                    "details": str(exc),
                }
            )
            candidate_counts[pair_label] = 0
            continue
        if evaluation_mode == STATION_MODE_ALL:
            pair_candidates = all_station_candidates(pair, index1, index2, excluded)
        else:
            pair_candidates = [
                selection
                for selection in selections
                if selection.pair_label == pair_label
                and selection.selection_status == "confirmed"
            ]
        candidate_counts[pair_label] = len(pair_candidates)
        pair_inputs[pair_label] = (index1, index2, pair_candidates)

    total_candidates = sum(candidate_counts.values())
    completed_candidates = 0
    report_station_progress(completed_candidates, total_candidates)
    for pair_label in pair_labels:
        pair = pairs[pair_label]
        if pair_label not in pair_inputs:
            continue
        index1, index2, pair_candidates = pair_inputs[pair_label]
        for selection in pair_candidates:
            progress_status = "failed"
            exception_distance_degrees: float | None = None
            try:
                candidates1 = index1.get(selection.station_id, [])
                candidates2 = index2.get(selection.station_id, [])
                if not candidates1 or not candidates2:
                    missing = (
                        "both"
                        if not candidates1 and not candidates2
                        else ("event1" if not candidates1 else "event2")
                    )
                    exceptions.append(
                        {
                            "pair_label": pair_label,
                            "station_id": selection.station_id,
                            "stage": "trace_match",
                            "exception": "missing_bhz_member",
                            "details": missing,
                        }
                    )
                    continue
                path1, path2 = choose_trace_pair(
                    candidates1,
                    candidates2,
                    float(config["station_coordinate_tolerance_degrees"]),
                )
                matched1 = next(row for row in candidates1 if row["path"] == path1)
                matched2 = next(row for row in candidates2 if row["path"] == path2)
                exception_distance_degrees = paired_epicentral_distance_degrees(
                    pair,
                    float(matched1["latitude"]),
                    float(matched1["longitude"]),
                    float(matched2["latitude"]),
                    float(matched2["longitude"]),
                )
                waveform_fingerprints_before = {
                    str(path): file_fingerprint(path) for path in (path1, path2)
                }
                used_paths.update((path1, path2))
                station = analyze_station(selection, pair, path1, path2, model, config)
                station_analyses[pair_label].append(station)
                noise_rows.extend(station.noise_rows)
                if station.result["new_correlation"] >= correlation_threshold:
                    plot_station_diagnostic(
                        station,
                        output
                        / "station_diagnostics"
                        / f"{pair_label}_{selection.network}_{selection.station}.png",
                    )
                for path, fingerprint in waveform_fingerprints_before.items():
                    fixed_fingerprints_before[path] = fingerprint
                progress_status = "completed"
            except Exception as exc:
                exceptions.append(
                    {
                        "pair_label": pair_label,
                        "station_id": selection.station_id,
                        "epicentral_distance_degrees": exception_distance_degrees,
                        "stage": "analysis",
                        "exception": type(exc).__name__,
                        "details": str(exc),
                    }
                )
            finally:
                completed_candidates += 1
                report_station_progress(
                    completed_candidates,
                    total_candidates,
                    pair_label,
                    selection.station_id,
                    progress_status,
                )

    station_rows = [
        station.result for label in pair_labels for station in station_analyses[label]
    ]
    threshold_rows = rows_meeting_correlation_threshold(
        station_rows, correlation_threshold
    )
    pair_rows: list[dict[str, Any]] = []
    for pair_offset, pair_label in enumerate(pair_labels):
        pair = pairs[pair_label]
        stations = station_analyses[pair_label]
        selected_count = sum(
            selection.pair_label == pair_label
            and selection.selection_status == "confirmed"
            for selection in selections
        )
        bootstrap = pair_bootstrap(
            stations,
            int(config["bootstrap_iterations"]),
            int(config["random_seed"]) + pair_offset,
            int(config["minimum_pair_stations"]),
        )
        r_values = np.asarray(
            [station.result["r_sym"] for station in stations], dtype=float
        )
        correlations = np.asarray(
            [station.result["new_correlation"] for station in stations], dtype=float
        )
        pair_exceptions = [row for row in exceptions if row["pair_label"] == pair_label]
        pair_result = {
            "pair_label": pair_label,
            "event1": pair.event1.event_id,
            "event2": pair.event2.event_id,
            "station_evaluation_mode": evaluation_mode,
            "candidate_station_count": candidate_counts.get(pair_label, 0),
            "selected_station_count": selected_count,
            "station_count": len(stations),
            "correlation_threshold": correlation_threshold,
            "correlation_threshold_station_count": sum(
                station.result["new_correlation"] >= correlation_threshold
                for station in stations
            ),
            **bootstrap,
            "median_r_sym": float(np.nanmedian(r_values))
            if len(r_values)
            else math.nan,
            "min_r_sym": float(np.nanmin(r_values)) if len(r_values) else math.nan,
            "max_r_sym": float(np.nanmax(r_values)) if len(r_values) else math.nan,
            "median_new_correlation": float(np.nanmedian(correlations))
            if len(correlations)
            else math.nan,
            "exception_count": len(pair_exceptions),
            "qc_flagged_station_count": sum(
                bool(station.result["qc_flags"]) for station in stations
            ),
        }
        pair_rows.append(pair_result)
        plotted_stations = stations_meeting_correlation_threshold(
            stations, correlation_threshold
        )
        plot_pair_overlay(
            pair,
            plotted_stations,
            output / "pair_plots" / f"{pair_label}.png",
            correlation_threshold,
        )
        plot_pair_summary(pair_result, output / "pair_summaries" / f"{pair_label}.png")

    write_csv_rows(output / "station_results.csv", station_rows)
    write_csv_rows(
        output / "stations_meeting_correlation_threshold.csv",
        threshold_rows,
        ["selection_correlation_threshold", *station_rows[0].keys()]
        if station_rows
        else ["selection_correlation_threshold"],
    )
    write_csv_rows(output / "pair_results.csv", pair_rows)
    write_csv_rows(
        output / "noise_chunks.csv",
        noise_rows,
        [
            "pair_label",
            "station_id",
            "event_number",
            "event_id",
            "chunk_index",
            "rms_stored_amplitude_units",
            "normalized_rms",
            "unusually_large",
        ],
    )
    write_csv_rows(
        output / "exceptions.csv",
        exceptions,
        [
            "pair_label",
            "station_id",
            "epicentral_distance_degrees",
            "stage",
            "exception",
            "details",
        ],
    )
    expected_outputs = [
        output / "analysis_config.json",
        output / "station_selection.csv",
        output / "station_results.csv",
        output / "stations_meeting_correlation_threshold.csv",
        output / "pair_results.csv",
        output / "noise_chunks.csv",
        output / "exceptions.csv",
        *[output / "pair_plots" / f"{label}.png" for label in pair_labels],
        *[output / "pair_summaries" / f"{label}.png" for label in pair_labels],
    ]
    missing_outputs = [str(path) for path in expected_outputs if not path.is_file()]
    if missing_outputs:
        raise AnalysisError(
            f"Expected outputs were not created: {', '.join(missing_outputs)}"
        )
    fingerprints_after = [
        file_fingerprint(path) for path in sorted(used_paths) if path.is_file()
    ]
    unchanged = all(
        fixed_fingerprints_before.get(fingerprint["path"]) == fingerprint
        for fingerprint in fingerprints_after
    )
    finished = utc_now_iso()
    elapsed_time_seconds = time.perf_counter() - elapsed_started
    cpu_time_seconds = time.process_time() - cpu_started
    manifest = {
        "run_started_utc": started,
        "run_finished_utc": finished,
        "elapsed_time_seconds": elapsed_time_seconds,
        "cpu_time_seconds": cpu_time_seconds,
        "output_directory": str(output),
        "software_versions": software_versions(),
        "random_seed": int(config["random_seed"]),
        "processing_parameters": config,
        "input_fingerprints": fingerprints_after,
        "source_inputs_unchanged_during_final_check": unchanged,
        "expected_outputs_verified": True,
        "counts": {
            "pairs_requested": len(pair_labels),
            "station_evaluation_mode": evaluation_mode,
            "station_candidates": sum(candidate_counts.values()),
            "confirmed_station_selections": sum(
                s.selection_status == "confirmed" for s in selections
            ),
            "station_results": len(station_rows),
            "stations_meeting_correlation_threshold": len(threshold_rows),
            "exceptions": len(exceptions),
        },
        "exclusions": exceptions,
    }
    write_strict_json(output / "run_manifest.json", manifest)
    return AnalysisRun(output, elapsed_time_seconds, cpu_time_seconds)


def build_parser() -> argparse.ArgumentParser:
    script_dir = Path(__file__).resolve().parent
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--config", type=Path, default=script_dir / "analysis_config.json"
    )
    parser.add_argument(
        "--selection", type=Path, default=script_dir / "station_selection.csv"
    )
    parser.add_argument(
        "--validate-only",
        action="store_true",
        help="Validate configuration, catalog pairs, and station-selection rows without reading SAC samples or writing outputs.",
    )
    return parser


def main(argv: Sequence[str] | None = None) -> int:
    elapsed_started = time.perf_counter()
    cpu_started = time.process_time()
    arguments = build_parser().parse_args(argv)
    try:
        config = load_json(arguments.config)
        validate_config_paths(config)
        evaluation_mode = station_evaluation_mode(config)
        excluded = excluded_station_codes(config)
        if arguments.validate_only:
            pairs = resolve_catalog(
                Path(config["catalog_path"]),
                config["pairs"],
                float(config["coordinate_tolerance_degrees"]),
                float(config["coordinate_tolerance_depth_km"]),
            )
            selections = [
                selection
                for selection in load_selections(arguments.selection)
                if not station_is_excluded(selection.station_id, excluded)
            ]
            validate_selections(
                selections, pairs, float(config["selection_correlation_threshold"])
            )
            if evaluation_mode == STATION_MODE_SELECTED:
                validate_selection_figures(selections, Path(config["figure_root"]))
            print(
                f"Validated {len(pairs)} pairs, "
                f"{sum(s.selection_status == 'confirmed' for s in selections)} confirmed selections, "
                f"and {sum(s.selection_status == 'review' for s in selections)} review items; "
                f"station evaluation mode is {evaluation_mode}."
            )
            report_runtime(
                time.perf_counter() - elapsed_started,
                time.process_time() - cpu_started,
            )
            return 0
        run = run_analysis(arguments.config, arguments.selection)
        print(run.output_directory)
        report_runtime(run.elapsed_time_seconds, run.cpu_time_seconds)
        return 0
    except Exception as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        if not isinstance(exc, AnalysisError):
            traceback.print_exc()
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
