#!/usr/bin/env python3
"""Compact pair relative-location columns in ICevents_full.xlsx."""

from __future__ import annotations

import argparse
import shutil
from datetime import datetime
from pathlib import Path
from typing import Any

import openpyxl


DEFAULT_WORKBOOK = Path("/Users/jvidale/Documents/GitHub/Array_codes/Files/ICevents_full.xlsx")

NEW_DELTA_COLUMNS = [
    "event2_minus_event1_delta_lat",
    "event2_minus_event1_delta_lon",
    "event2_minus_event1_delta_depth_km",
    "event2_minus_event1_delta_time_s",
]

REMOVE_COLUMNS = [
    "event1_delta_lat",
    "event1_delta_lon",
    "event1_delta_depth_km",
    "event1_delta_time_s",
    "event2_delta_lat",
    "event2_delta_lon",
    "event2_delta_depth_km",
    "event2_delta_time_s",
    "event2_minus_event1_origin_time_shift_s",
    "event2_minus_event1_east_km",
    "event2_minus_event1_north_km",
    "event2_minus_event1_horizontal_km",
]


def header_map(sheet) -> dict[str, int]:
    return {
        str(cell.value).strip().lower(): cell.column
        for cell in sheet[1]
        if cell.value is not None and str(cell.value).strip()
    }


def cell_value(row, header: dict[str, int], name: str) -> Any:
    column = header.get(name.lower())
    if column is None:
        return None
    return row[column - 1]


def finite_float(value: Any) -> float | None:
    try:
        parsed = float(value)
    except (TypeError, ValueError):
        return None
    return parsed


def compact_workbook(workbook_path: Path) -> Path:
    backup_path = (
        workbook_path.parent
        / f"{workbook_path.stem}_before_compact_pair_relative_columns_"
        f"{datetime.now().strftime('%Y%m%d_%H%M%S')}{workbook_path.suffix}"
    )
    shutil.copy2(workbook_path, backup_path)

    workbook = openpyxl.load_workbook(workbook_path)
    try:
        sheet = workbook["pairs"]
        header = header_map(sheet)
        insert_after = header["centroid_depth_km"]
        for name in reversed(NEW_DELTA_COLUMNS):
            header = header_map(sheet)
            if name.lower() in header:
                continue
            sheet.insert_cols(insert_after + 1)
            sheet.cell(row=1, column=insert_after + 1).value = name

        header = header_map(sheet)
        written = 0
        for row_index in range(2, sheet.max_row + 1):
            values = [
                sheet.cell(row=row_index, column=column).value
                for column in range(1, sheet.max_column + 1)
            ]
            event2_delta_lat = finite_float(
                cell_value(values, header, "event2_delta_lat")
            )
            event2_delta_lon = finite_float(
                cell_value(values, header, "event2_delta_lon")
            )
            event2_delta_depth = finite_float(
                cell_value(values, header, "event2_delta_depth_km")
            )
            event2_delta_time = finite_float(
                cell_value(values, header, "event2_delta_time_s")
            )
            origin_shift = finite_float(
                cell_value(values, header, "event2_minus_event1_origin_time_shift_s")
            )
            new_values = {
                "event2_minus_event1_delta_lat": None
                if event2_delta_lat is None
                else 2.0 * event2_delta_lat,
                "event2_minus_event1_delta_lon": None
                if event2_delta_lon is None
                else 2.0 * event2_delta_lon,
                "event2_minus_event1_delta_depth_km": None
                if event2_delta_depth is None
                else 2.0 * event2_delta_depth,
                "event2_minus_event1_delta_time_s": origin_shift
                if origin_shift is not None
                else (None if event2_delta_time is None else 2.0 * event2_delta_time),
            }
            if all(value is None for value in new_values.values()):
                continue
            for name, value in new_values.items():
                cell = sheet.cell(row=row_index, column=header[name.lower()])
                cell.value = value
                cell.number_format = "0.0000"
            written += 1

        header = header_map(sheet)
        removable = sorted(
            [header[name.lower()] for name in REMOVE_COLUMNS if name.lower() in header],
            reverse=True,
        )
        for column in removable:
            sheet.delete_cols(column)

        workbook.save(workbook_path)
    finally:
        workbook.close()

    print(f"backup={backup_path}")
    print(f"written={written}")
    print(f"removed_columns={len(removable)}")
    return backup_path


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--workbook", type=Path, default=DEFAULT_WORKBOOK)
    args = parser.parse_args()
    compact_workbook(args.workbook)


if __name__ == "__main__":
    main()
