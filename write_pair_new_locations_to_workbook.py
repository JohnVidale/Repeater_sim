#!/usr/bin/env python3
"""Write median fixed-depth pair locations into ICevents_full.xlsx."""

from __future__ import annotations

import argparse
import csv
import shutil
from datetime import datetime
from pathlib import Path
from typing import Any

import openpyxl


def normalized_header(row: tuple[Any, ...]) -> dict[str, int]:
    return {
        str(value).strip().lower(): index + 1
        for index, value in enumerate(row)
        if value is not None and str(value).strip()
    }


def read_summary(path: Path) -> dict[str, tuple[float, float]]:
    values: dict[str, tuple[float, float]] = {}
    with path.open(newline="", encoding="utf-8") as handle:
        for row in csv.DictReader(handle):
            if row.get("status") != "ok":
                continue
            label = str(row["pair_label"]).strip()
            values[label] = (float(row["new_lat"]), float(row["new_lon"]))
    return values


def ensure_column(sheet, header: dict[str, int], name: str, after_column: int) -> int:
    key = name.strip().lower()
    if key in header:
        return header[key]
    sheet.insert_cols(after_column + 1)
    sheet.cell(row=1, column=after_column + 1).value = name
    return after_column + 1


def write_locations(workbook_path: Path, summary_path: Path) -> Path:
    locations = read_summary(summary_path)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_path = (
        workbook_path.parent
        / f"{workbook_path.stem}_before_pair_new_lat_lon_{stamp}{workbook_path.suffix}"
    )
    shutil.copy2(workbook_path, backup_path)

    workbook = openpyxl.load_workbook(workbook_path)
    try:
        sheet = workbook["pairs"]
        header = normalized_header(
            tuple(cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1)))
        )
        label_column = header["label"]
        lat_column = header["lat"]
        lon_column = header["lon"]
        new_lat_column = ensure_column(sheet, header, "new_lat", lat_column)
        if new_lat_column <= lon_column:
            lon_column += 1
        header = normalized_header(
            tuple(cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1)))
        )
        new_lon_column = ensure_column(sheet, header, "new_lon", lon_column)

        written = 0
        for row_index in range(2, sheet.max_row + 1):
            label_value = sheet.cell(row=row_index, column=label_column).value
            if label_value is None:
                continue
            label = str(label_value).strip()
            if label not in locations:
                continue
            latitude, longitude = locations[label]
            lat_cell = sheet.cell(row=row_index, column=new_lat_column)
            lon_cell = sheet.cell(row=row_index, column=new_lon_column)
            lat_cell.value = latitude
            lon_cell.value = longitude
            lat_cell.number_format = "0.0000"
            lon_cell.number_format = "0.0000"
            written += 1

        workbook.save(workbook_path)
    finally:
        workbook.close()
    print(f"backup={backup_path}")
    print(f"written={written}")
    return backup_path


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("workbook", type=Path)
    parser.add_argument("summary", type=Path)
    arguments = parser.parse_args()
    write_locations(arguments.workbook, arguments.summary)


if __name__ == "__main__":
    main()
