#!/usr/bin/env python3
"""Plot repeater-pair age against the fitted 3-D event separation.

The plot combines the preferred relative-location fit (excluding PKiKP) with
the median direct-P correlation for each pair.  Pair groups are read from the
``pairs_purple`` and ``pairs_blue`` lists in the analysis JSON file.
"""

from __future__ import annotations

import argparse
import csv
import json
from datetime import datetime
from pathlib import Path

import matplotlib.pyplot as plt
from matplotlib.lines import Line2D


ROOT = Path(__file__).resolve().parent
DEFAULT_CONFIG = ROOT / "analysis_config.json"
DEFAULT_OUTPUTS = ROOT / "outputs"
LOCATION_FILE = "median_relative_location_offsets_no_pkikp_with_bootstrap_uncertainty.csv"
PHASE_FILE = "phase_summary.csv"


def read_csv(path: Path) -> list[dict[str, str]]:
    with path.open(newline="") as handle:
        return list(csv.DictReader(handle))


def newest_complete_output(outputs: Path) -> Path:
    candidates = [
        path
        for path in outputs.glob("multiphase_median_*")
        if (path / LOCATION_FILE).is_file() and (path / PHASE_FILE).is_file()
    ]
    if not candidates:
        raise FileNotFoundError(f"No multiphase output containing {LOCATION_FILE} found in {outputs}")
    return max(candidates, key=lambda path: path.stat().st_mtime)


def parse_date(value: str) -> datetime:
    return datetime.fromisoformat(str(value).strip()).replace(tzinfo=None)


def main() -> Path:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--config", type=Path, default=DEFAULT_CONFIG)
    parser.add_argument(
        "--output-directory",
        type=Path,
        help="Multiphase output directory. Defaults to the newest complete one.",
    )
    parser.add_argument("--output", type=Path, help="PNG path (default: in the output directory).")
    args = parser.parse_args()

    config = json.loads(args.config.read_text())
    output_directory = args.output_directory or newest_complete_output(DEFAULT_OUTPUTS)
    output_path = args.output or output_directory / "pair_age_vs_relative_separation.png"

    purple = set(map(str, config.get("pairs_purple", [])))
    blue = set(map(str, config.get("pairs_blue", [])))
    requested = [str(pair) for pair in config["pairs"]]

    locations = {row["pair"]: row for row in read_csv(output_directory / LOCATION_FILE)}
    p_cc = {
        row["pair_label"]: float(row["median_cc"])
        for row in read_csv(output_directory / PHASE_FILE)
        if row["phase"] == "P" and row["median_cc"]
    }

    # The workbook is the authoritative source for pair event dates.
    workbook_path = Path(config.get("time_shift_workbook") or config["catalog_path"])
    from openpyxl import load_workbook

    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    sheet = workbook["pairs"]
    headers = [cell.value for cell in next(sheet.iter_rows(max_row=1))]
    pair_dates = {
        str(row[headers.index("label")]): (row[headers.index("date1")], row[headers.index("date2")])
        for row in sheet.iter_rows(min_row=2, values_only=True)
        if row[headers.index("label")] is not None
    }
    workbook.close()

    plotted: list[str] = []
    missing: list[str] = []
    figure, axis = plt.subplots(figsize=(12.5, 7.5), constrained_layout=True)
    colors = {"purple": "#6a3d9a", "blue": "#1f78b4", "other": "#555555"}

    for index, pair in enumerate(requested):
        row = locations.get(pair)
        dates = pair_dates.get(pair)
        if row is None or dates is None or pair not in p_cc:
            missing.append(pair)
            continue
        date1, date2 = dates
        years = abs((parse_date(str(date2)) - parse_date(str(date1))).days) / 365.25
        separation = float(row["separation_3d_km"])
        group = "purple" if pair in purple else "blue" if pair in blue else "other"
        color = colors[group]
        axis.scatter(years, separation, s=86, color=color, edgecolor="black", linewidth=0.6, zorder=3)
        # Hand-tuned offsets keep the dense low-separation cluster legible
        # while retaining the same labels and data coordinates.
        label_offsets = {
            "P30": (6, 12), "P31": (8, 10), "P33": (8, 10),
            "P34": (10, 14), "P35": (10, -15), "P37": (8, -15),
            "P38": (8, 12), "P39": (8, -15), "P51": (-62, 18),
            "P79": (8, 12), "P112": (-68, 10), "P123": (8, -15),
            "P140": (-72, 25), "P145": (-72, -15), "P321": (-70, -15),
            "P355": (8, 10),
        }
        horizontal, vertical = label_offsets.get(pair, (6, 10))
        axis.annotate(
            f"{pair}  P cc={p_cc[pair]:.2f}",
            (years, separation),
            xytext=(horizontal, vertical),
            textcoords="offset points",
            fontsize=9,
            ha="right" if horizontal < 0 else "left",
            va="bottom" if vertical > 0 else "top",
        )
        plotted.append(pair)

    axis.set_xlabel("Time between events (years)")
    axis.set_ylabel("Preferred 3-D repeater separation (km)")
    axis.set_title("Repeater-pair interval and relative separation\n(labels give pair and median P-wave correlation)")
    axis.grid(True, alpha=0.28, zorder=0)
    axis.set_ylim(0, max(1.82, axis.get_ylim()[1]))
    axis.legend(
        handles=[
            Line2D([], [], marker="o", linestyle="", markersize=11, markerfacecolor=colors["purple"], markeredgecolor="black", label="Purple group"),
            Line2D([], [], marker="o", linestyle="", markersize=11, markerfacecolor=colors["blue"], markeredgecolor="black", label="Blue group"),
        ],
        loc="upper left",
        frameon=True,
    )
    figure.savefig(output_path, dpi=220)
    plt.close(figure)

    print(output_path)
    print(f"Plotted {len(plotted)} pairs: {', '.join(plotted)}")
    if missing:
        print(f"Not plotted (missing fit, date, or P correlation): {', '.join(missing)}")
    return output_path


if __name__ == "__main__":
    main()
